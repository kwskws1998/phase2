import argparse
import os
import json
import sys
import threading
import subprocess
import tempfile

# Early GPU Configuration (MUST be before torch import)
from utils.gpu_config import configure_gpu
configure_gpu()

import torch
import torch.nn as nn
import webbrowser
import re
import base64
import io
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn
from urllib.parse import urlparse
from transformers import AutoTokenizer, AutoProcessor, TextStreamer
import model as model_module
import traceback
from utils import get_file_config
from utils.paths import set_local_mode, get_model_dir, get_log_dir, ensure_dirs

# Tool system imports
try:
    from tools import execute_tool_call, parse_tool_calls
    from tools.executor import format_tool_result_for_llm, set_adapter, detect_tool_call
    from tools.base import get_tools_schema
    # Import adapters
    from tools.adapters import get_adapter_for_model, ToolResult
    # Import to register tools
    import tools.biomni.bio_tools
    import tools.plan.plan_tools
    import tools.code.code_tools
    import tools.analysis.analysis_tools
    TOOLS_AVAILABLE = True
except ImportError as e:
    print(f"[Warning] Tool system not available: {e}")
    TOOLS_AVAILABLE = False
    
    # Fallback stubs
    def set_adapter(model_type): pass
    def detect_tool_call(text): return False
    def get_tools_schema(): return []
    def get_adapter_for_model(model_type): return None
    class ToolResult: pass

# Optional: PIL for image processing
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# Threading HTTP Server
class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True

# MODEL_BASE_DIR is set dynamically based on --local flag
MODEL_BASE_DIR = None  # Will be set in main()

# Retry configuration - unified for all retry operations
MAX_RETRY_ATTEMPTS = 3

# Default system prompt if SYSTEM_PROMPT.txt not found
DEFAULT_SYSTEM_PROMPT = "You are a helpful AI assistant."

# Meta tokens to filter (structural, not semantic)
# These are always auto-added and have no semantic meaning in output
# <unk>=0, <s>=1, </s>=2, <pad>=11
META_TOKEN_IDS = {0, 1, 2, 11}
META_TOKEN_STRINGS = {'<s>', '</s>', '<pad>', '<unk>'}

def strip_meta_tokens(text: str) -> str:
    """Remove only structural meta tokens from string.
    
    Tool-related tokens like [TOOL_CALLS], [ARGS] are preserved.
    """
    for token in META_TOKEN_STRINGS:
        text = text.replace(token, '')
    return text


from utils.model_utils import get_model_device, get_unwrapped_model


def calculate_uncertainty(logits, vocab_size):
    """
    Calculate uncertainty as standard deviation of logits.
    
    This is consistent with the heteroscedastic head used in training,
    where sigma = std(logits) represents model uncertainty.
    
    Args:
        logits: Token logits tensor of shape (batch, vocab_size)
        vocab_size: Size of vocabulary (unused, kept for API compatibility)
    
    Returns:
        float: Uncertainty score (std of logits)
               Higher values = more spread in logit values = more uncertainty
    """
    # σ = std(logits) - same as heteroscedastic head in loss.py
    sigma = logits.std(dim=-1)
    return sigma.item()


def sample_token(logits, temperature, top_k):
    """
    Sample a token from logits with temperature and top-k sampling.
    
    Args:
        logits: Token logits tensor of shape (batch, vocab_size)
        temperature: Sampling temperature (higher = more random)
        top_k: Number of top tokens to consider
    
    Returns:
        torch.Tensor: Sampled token id
    """
    # Apply temperature
    if temperature > 0:
        logits = logits / temperature
    
    # Apply top-k filtering
    if top_k > 0:
        top_k = min(top_k, logits.size(-1))
        indices_to_remove = logits < torch.topk(logits, top_k)[0][..., -1, None]
        logits[indices_to_remove] = float('-inf')
    
    # Sample from distribution
    probs = torch.softmax(logits, dim=-1)
    next_token = torch.multinomial(probs, num_samples=1)
    
    return next_token.squeeze(-1)


def compute_recovery_temp(lowest, target, step, total_steps, method="linear"):
    """
    Calculate temperature recovery after refusal.
    
    When a token is refused due to high uncertainty, the temperature is lowered.
    This function computes the gradual recovery back to the original temperature.
    
    Args:
        lowest: The lowest temperature reached after refusal
        target: The target (original) temperature to recover to
        step: Current step in recovery (0 = just after refusal)
        total_steps: Total steps to fully recover
        method: Recovery curve type
            - "linear": Constant recovery rate
            - "exponential": Slow start, fast end (ease-in)
            - "ease_out": Fast start, slow end
            - "ease_in_out": S-curve (slow-fast-slow)
            - "step": Stay at lowest until total_steps, then jump to target
    
    Returns:
        float: Current effective temperature
    """
    if step >= total_steps:
        return target
    
    if total_steps <= 0:
        return target
    
    progress = step / total_steps  # 0.0 ~ 1.0
    gap = target - lowest
    
    if method == "linear":
        return lowest + gap * progress
    elif method == "exponential":
        # Ease-in: slow start, accelerating
        return lowest + gap * (progress ** 1.5)
    elif method == "ease_out":
        # Fast start, decelerating
        return lowest + gap * (1 - (1 - progress) ** 2)
    elif method == "ease_in_out":
        # S-curve: slow-fast-slow
        if progress < 0.5:
            return lowest + gap * (2 * progress ** 2)
        else:
            return lowest + gap * (1 - (-2 * progress + 2) ** 2 / 2)
    elif method == "step":
        # Stay at lowest until fully recovered
        return lowest
    else:
        # Default to linear
        return lowest + gap * progress


def generate_with_refusal_streaming(model, tokenizer, inputs, args):
    """
    Generator version of token-by-token generation with refusal mechanism.
    
    Yields tokens one by one for streaming output (SSE, etc.).
    Includes temperature recovery after refusal.
    
    Args:
        model: The model to generate from
        tokenizer: Tokenizer for decoding
        inputs: Tokenized inputs (dict with input_ids, attention_mask)
        args: Generation arguments including:
            - temperature: Base sampling temperature
            - max_length: Maximum tokens to generate
            - top_k: Top-k sampling parameter
            - refusal_threshold: Uncertainty threshold for refusal
            - refusal_max_retries: Max retries per token
            - refusal_temp_decay: Temperature multiplier on refusal
            - refusal_recovery_tokens: Tokens to recover to original temp
            - refusal_recovery_method: Recovery curve type
    
    Yields:
        dict: {"token": str, "done": bool, "full_text": str (only when done)}
    """
    import random as py_random
    
    # Handle both dict-like and tensor inputs
    # Note: We don't use attention_mask for single-sequence inference (no padding)
    # This allows SDPA to use is_causal=True internally, avoiding O(n²) mask creation
    device = get_model_device(model)
    if hasattr(inputs, 'input_ids'):
        input_ids = inputs.input_ids.to(device)
    else:
        input_ids = inputs['input_ids'].to(device)
    
    # Extract pixel_values for multimodal input
    pixel_values = None
    if hasattr(inputs, 'pixel_values') and inputs.pixel_values is not None:
        pixel_values = inputs.pixel_values.to(device)
    elif isinstance(inputs, dict) and 'pixel_values' in inputs and inputs['pixel_values'] is not None:
        pixel_values = inputs['pixel_values'].to(device)
    
    vocab_size = get_unwrapped_model(model).config.vocab_size
    
    # Get refusal parameters with defaults
    threshold = getattr(args, 'refusal_threshold', 3.0)
    max_retries = getattr(args, 'refusal_max_retries', 3)
    temp_decay = getattr(args, 'refusal_temp_decay', 0.8)
    min_temp = getattr(args, 'refusal_min_temp', 0.4)
    random_seed = getattr(args, 'random_seed', -1)
    use_random_seed = random_seed < 0
    debug = getattr(args, 'debug', False)
    
    # Recovery parameters
    recovery_tokens = getattr(args, 'refusal_recovery_tokens', 3)
    recovery_method = getattr(args, 'refusal_recovery_method', 'exponential')
    
    # Temperature state for recovery
    target_temp = args.temperature
    effective_temp = target_temp
    lowest_temp = target_temp
    tokens_since_refusal = float('inf')  # inf = not in recovery mode
    min_temp_used = target_temp  # Track minimum temperature actually used
    
    generated_tokens = []
    generated_text = ""
    consecutive_pad_count = 0
    pending_token_ids = []  # UTF-8 byte token buffer
    temp_sum = 0.0  # For avg_temp calculation
    temp_count = 0
    
    # Memory logging
    import torch.cuda as cuda
    print(f"\n{'='*50}")
    print(f"[Memory] Generation Start")
    print(f"[Memory] Input tokens: {input_ids.shape[1]}")
    print(f"[Memory] VRAM allocated: {cuda.memory_allocated() / 1024**3:.2f} GB")
    print(f"[Memory] VRAM reserved: {cuda.memory_reserved() / 1024**3:.2f} GB")
    
    # Initial forward pass: compute KV cache for the entire input sequence
    # Include pixel_values for multimodal models (image processing happens here)
    # Note: attention_mask=None allows SDPA to use is_causal=True (more efficient)
    seq_len = input_ids.shape[1]
    position_ids = torch.arange(seq_len, device=device).unsqueeze(0)
    
    with torch.no_grad():
        outputs = model(
            input_ids, 
            attention_mask=None,  # No padding, let SDPA handle causal masking
            position_ids=position_ids,
            pixel_values=pixel_values,  # Pass image data for vision encoder
            use_cache=True
        )
    past_key_values = outputs.past_key_values
    next_logits = outputs.logits[:, -1, :]
    
    print(f"[Memory] After initial forward: {cuda.memory_allocated() / 1024**3:.2f} GB")
    print(f"[Memory] KV cache layers: {len(past_key_values)}")
    if past_key_values and past_key_values[0]:
        print(f"[Memory] KV cache shape: {past_key_values[0][0].shape}")
    
    # After initial forward, pixel_values is no longer needed (image features are cached)
    pixel_values = None
    
    # Track position for RoPE
    current_position = input_ids.shape[1]
    
    for step in range(args.max_length):
        if global_stop_generation:
            break
        
        # Memory logging every 500 tokens
        if step > 0 and step % 500 == 0:
            kv_seq_len = past_key_values[0][0].shape[2] if past_key_values else 0
            print(f"[Memory] Step {step}: Generated {len(generated_tokens)} tokens, "
                  f"Total context: {kv_seq_len}, "
                  f"VRAM: {cuda.memory_allocated() / 1024**3:.2f} GB")
        
        # Temperature recovery logic
        if tokens_since_refusal < recovery_tokens:
            effective_temp = compute_recovery_temp(
                lowest_temp, target_temp, 
                tokens_since_refusal, recovery_tokens, 
                recovery_method
            )
            tokens_since_refusal += 1
        else:
            effective_temp = target_temp
        
        current_temp = effective_temp
        was_refused = False
        
        # Look-ahead retry loop: sample token, forward pass, check future uncertainty
        for retry in range(max_retries + 1):
            # Set seed for sampling
            if use_random_seed:
                current_seed = py_random.randint(0, 2**32 - 1)
            else:
                current_seed = random_seed + step * 100 + retry
            torch.manual_seed(current_seed)
            if torch.cuda.is_available():
                torch.cuda.manual_seed(current_seed)
            
            # 1. Sample candidate token from current logits
            candidate_token = sample_token(next_logits, current_temp, args.top_k)
            
            # 2. Forward pass with candidate to get future logits
            # Note: With KV cache, we only process 1 token, position_ids handles RoPE
            step_position_ids = torch.tensor([[current_position]], device=device)
            
            with torch.no_grad():
                future_outputs = model(
                    candidate_token.view(1, 1),
                    attention_mask=None,  # No padding, SDPA handles causal masking
                    position_ids=step_position_ids,
                    past_key_values=past_key_values,
                    use_cache=True
                )
            future_logits = future_outputs.logits[:, -1, :]
            
            # 3. Check uncertainty of the FUTURE logits
            uncertainty = calculate_uncertainty(future_logits, vocab_size)
            
            if debug:
                candidate_text = tokenizer.decode(candidate_token, skip_special_tokens=True)
                # Handle Windows encoding issues
                try:
                    print(f"\n[DEBUG] Token {len(generated_tokens)}, retry {retry}: "
                          f"candidate='{candidate_text}', future_uncertainty={uncertainty:.4f}, temp={current_temp:.3f}", end="")
                except UnicodeEncodeError:
                    safe_text = candidate_text.encode('ascii', errors='replace').decode('ascii')
                    print(f"\n[DEBUG] Token {len(generated_tokens)}, retry {retry}: "
                          f"candidate='{safe_text}', future_uncertainty={uncertainty:.4f}, temp={current_temp:.3f}", end="")
            
            # Accept if: uncertainty is low, max retries reached, or temp already at minimum
            at_min_temp = current_temp <= min_temp
            if uncertainty <= threshold or retry == max_retries or at_min_temp:
                # Accept: update state with this candidate
                next_token = candidate_token
                past_key_values = future_outputs.past_key_values
                next_logits = future_logits  # Use for next step
                current_position += 1
                if debug and retry > 0:
                    if at_min_temp:
                        print(f" -> ACCEPTED (min temp reached)")
                    else:
                        print(f" -> ACCEPTED after {retry} retries")
                elif debug:
                    print(f" -> ACCEPTED")
                break
            else:
                # Refuse: lower temperature and retry (don't update KV cache)
                was_refused = True
                if debug:
                    print(f" -> REFUSED, retry {retry+1}/{max_retries}")
                current_temp *= temp_decay
                current_temp = max(current_temp, min_temp)
                min_temp_used = min(min_temp_used, current_temp)  # Track minimum
        
        # If refused at least once, start recovery from the lowered temperature
        if was_refused:
            lowest_temp = current_temp
            tokens_since_refusal = 0
        
        # Decode token with UTF-8 byte buffering
        token_id = next_token.item()
        # Filter meta tokens at token ID level, preserve tool tokens
        if token_id not in META_TOKEN_IDS:
            token_text = tokenizer.decode([token_id], skip_special_tokens=False)
        else:
            token_text = ""  # meta token - skip output
        generated_tokens.append(token_id)
        
        if '\ufffd' in token_text:
            # Incomplete UTF-8 byte token → buffer it
            pending_token_ids.append(token_id)
            if debug:
                vocab_token = tokenizer.convert_ids_to_tokens([token_id])[0]
                try:
                    print(f"\n[DEBUG] Buffering incomplete UTF-8 token: {vocab_token}")
                except UnicodeEncodeError:
                    safe_token = str(vocab_token).encode('ascii', errors='replace').decode('ascii')
                    print(f"\n[DEBUG] Buffering incomplete UTF-8 token: {safe_token}")
        else:
            # Complete token → check if we have buffered tokens
            if pending_token_ids:
                # Decode buffered tokens together with current token
                all_ids = pending_token_ids + [token_id]
                token_text = tokenizer.decode(all_ids, skip_special_tokens=False)
                token_text = strip_meta_tokens(token_text)
                pending_token_ids = []
            
            generated_text += token_text
            # Track temperature for avg calculation
            temp_sum += current_temp
            temp_count += 1
            # Yield token with current temp info
            yield {"token": token_text, "done": False, "current_temp": current_temp, "min_temp_used": min_temp_used}
        
        # Check for consecutive pad tokens
        if next_token.item() == tokenizer.pad_token_id:
            consecutive_pad_count += 1
            if consecutive_pad_count >= 10:
                if debug:
                    print(f"\n[DEBUG] Stopping: {consecutive_pad_count} consecutive pad tokens")
                break
        else:
            consecutive_pad_count = 0
        
        # Check for EOS
        if next_token.item() == tokenizer.eos_token_id:
            break
        
        # Note: Forward pass for next token is now done in the retry loop above
        # past_key_values, next_logits, current_position are already updated
    
    # Handle remaining buffered tokens
    if pending_token_ids:
        remaining_text = tokenizer.decode(pending_token_ids, skip_special_tokens=False)
        remaining_text = strip_meta_tokens(remaining_text)
        if remaining_text and '\ufffd' not in remaining_text:
            generated_text += remaining_text
            yield {"token": remaining_text, "done": False, "current_temp": target_temp, "min_temp_used": min_temp_used}
    
    # Calculate average temperature
    avg_temp_used = temp_sum / temp_count if temp_count > 0 else target_temp
    
    # Memory logging - generation complete
    print(f"\n[Memory] Generation Complete")
    print(f"[Memory] Input tokens: {input_ids.shape[1]}")
    print(f"[Memory] Output tokens: {len(generated_tokens)}")
    print(f"[Memory] Total context: {current_position}")
    print(f"[Memory] Peak VRAM: {cuda.max_memory_allocated() / 1024**3:.2f} GB")
    print(f"{'='*50}\n")
    cuda.reset_peak_memory_stats()
    
    # Signal completion
    yield {"token": "", "done": True, "full_text": generated_text.strip(), "min_temp_used": min_temp_used, "avg_temp_used": avg_temp_used}


def generate_with_refusal(model, tokenizer, inputs, args):
    """
    Token-by-token generation with refusal mechanism and KV Cache.
    
    Wrapper around generate_with_refusal_streaming that prints to console
    and returns the final string.
    
    Args:
        model: The model to generate from
        tokenizer: Tokenizer for decoding
        inputs: Tokenized inputs (dict with input_ids, attention_mask)
        args: Generation arguments including refusal parameters
    
    Returns:
        str: Generated response text
    """
    full_text = ""
    for chunk in generate_with_refusal_streaming(model, tokenizer, inputs, args):
        if not chunk.get("done", False):
            token = chunk.get("token", "")
            try:
                print(token, end="", flush=True)
            except UnicodeEncodeError:
                safe_token = token.encode('ascii', errors='replace').decode('ascii')
                print(safe_token, end="", flush=True)
            full_text += token
        else:
            full_text = chunk.get("full_text", full_text)
    
    return full_text.strip()


def has_model_files(path):
    """
    Check if a directory contains model files (config.json and weights).
    """
    if not os.path.isdir(path):
        return False
    
    has_config = os.path.exists(os.path.join(path, "config.json"))
    has_weights = any(
        f.endswith(".bin") or f.endswith(".safetensors")
        for f in os.listdir(path)
    )
    return has_config and has_weights


def find_model_path(model_name, base_dir=None):
    """
    Find model path from model name.
    Searches in model/ and model/train/ directories.
    """
    if base_dir is None:
        base_dir = MODEL_BASE_DIR
    # 1. model/{model_name}
    path1 = os.path.join(base_dir, model_name)
    if os.path.exists(path1) and has_model_files(path1):
        return path1
    
    # 2. model/train/{model_name}
    path2 = os.path.join(base_dir, "train", model_name)
    if os.path.exists(path2) and has_model_files(path2):
        return path2
    
    # 3. Direct path (for train/xxx format)
    if "/" in model_name or "\\" in model_name:
        path3 = os.path.join(base_dir, model_name)
        if os.path.exists(path3) and has_model_files(path3):
            return path3
    
    # 4. Not found locally — try auto-download from HuggingFace Hub
    from utils.model_download import auto_download_model
    result = auto_download_model(model_name, base_dir, has_model_files_fn=has_model_files)
    if result:
        return result
    
    return None


def list_available_models(base_dir=None):
    """
    List all available models in the model directory.
    """
    if base_dir is None:
        base_dir = MODEL_BASE_DIR
    models = []
    
    if not base_dir or not os.path.exists(base_dir):
        return models
    
    # model/ direct subdirectories (excluding train)
    for name in os.listdir(base_dir):
        path = os.path.join(base_dir, name)
        if os.path.isdir(path) and name != "train":
            if has_model_files(path):
                models.append(name)
    
    # model/train/ subdirectories
    train_dir = os.path.join(base_dir, "train")
    if os.path.exists(train_dir):
        for name in os.listdir(train_dir):
            path = os.path.join(train_dir, name)
            if os.path.isdir(path) and has_model_files(path):
                models.append(f"train/{name}")
    
    return sorted(models)


def print_available_models():
    """
    Print all available models and exit.
    """
    models = list_available_models()
    
    print("\nAvailable models:")
    print("="*50)
    
    if not models:
        print("  (No models found)")
    else:
        for m in models:
            print(f"  - {m}")
    
    print("="*50)
    print(f"\nUsage: python inference.py --model <model_name>")
    print()


def _build_ref_context(step):
    """Build reference context string from step's references."""
    refs = step.get('references', [])
    if not refs:
        return ""
    parts = ["\n\n[Reference Data]"]
    for ref in refs:
        if ref.get('nodeType') == 'data':
            pv = ref.get('portValues', {}).get('out') or {}
            if pv.get('textContent'):
                parts.append(f"- {ref.get('title', 'Data')}: {pv['textContent'][:2000]}")
            elif pv.get('uploadId'):
                parts.append(f"- {ref.get('title', 'Data')}: file at /uploads/{pv['uploadId']}")
        elif ref.get('nodeType') == 'image':
            pv = ref.get('portValues', {}).get('out') or {}
            if pv.get('uploadId'):
                parts.append(
                    f"- {ref.get('title', 'Image')}: Image file available. "
                    f"Use view_image tool with image_path=\"/uploads/{pv['uploadId']}\" to analyze it."
                )
        else:
            parts.append(f"- Refer to results from: {ref.get('title', '')}")
    return "\n".join(parts)


def _collect_file_refs(step):
    """Extract file paths from step references for code_gen file reading."""
    paths = []
    for ref in step.get('references', []):
        if ref.get('nodeType') == 'data':
            pv = ref.get('portValues', {}).get('out') or {}
            if pv.get('uploadId'):
                paths.append(f'/uploads/{pv["uploadId"]}')
    return paths


def _collect_step_images(step):
    """Collect PIL Images from step's flow-connected image inputs.
    
    Used during plan execution to detect if a step (e.g. Composite node)
    has image data flowing into it, enabling multimodal inference.
    """
    images = []
    for port_name, input_data in step.get('inputs', {}).items():
        if input_data.get('nodeType') == 'image':
            pv = input_data.get('portValues', {}).get('out') or {}
            upload_id = pv.get('uploadId')
            if upload_id:
                img_path = os.path.join(UPLOADS_DIR, upload_id)
                if os.path.isfile(img_path):
                    try:
                        images.append(Image.open(img_path).convert('RGB'))
                    except Exception as e:
                        print(f"[Warning] Failed to load image {upload_id}: {e}")
    return images


def _build_step_inputs(step, history, system_prompt, tokenizer, args, tools_schema):
    """Build model inputs for a plan step, using multimodal if images are present."""
    step_images = _collect_step_images(step)
    if step_images and global_processor is not None:
        print(f"[DEBUG] Multimodal step detected: {len(step_images)} image(s)")
        messages = build_multimodal_messages(history, "", step_images, system_prompt)
        return build_inputs_multimodal(messages, step_images, global_processor, tokenizer, args)
    else:
        messages = build_messages(history, "", system_prompt)
        return build_inputs(messages, tokenizer, args, tools=tools_schema)


def load_tool_select_prompt():
    """Load the TOOL_SELECT_PROMPT for selecting tools during plan execution."""
    from tools.base import generate_tools_format
    
    prompt_path = os.path.join(os.path.dirname(__file__), "prompts", "TOOL_SELECT_PROMPT.txt")
    if os.path.exists(prompt_path):
        with open(prompt_path, "r", encoding="utf-8") as f:
            prompt = f.read().strip()
            # Replace {TOOLS_FORMAT} placeholder with dynamic content
            if "{TOOLS_FORMAT}" in prompt:
                prompt = prompt.replace("{TOOLS_FORMAT}", generate_tools_format())
            return prompt
    return "You select and call tools. Output ONLY one tool call, no explanations."


def load_system_prompt(model_path, model_type="ministral_3_3b_instruct"):
    """
    Load system prompt from prompts folder (priority) or model directory.
    Falls back to default if not found.
    Uses FileConfig.SYSTEM_PROMPT for the filename.
    """
    file_config = get_file_config(model_type)
    system_prompt_file = file_config.SYSTEM_PROMPT if file_config and hasattr(file_config, 'SYSTEM_PROMPT') else "SYSTEM_PROMPT.txt"
    
    def _process_prompt(prompt):
        """Replace date placeholders in prompt."""
        today = datetime.now().strftime("%Y-%m-%d")
        yesterday = (datetime.now() - __import__('datetime').timedelta(days=1)).strftime("%Y-%m-%d")
        prompt = prompt.replace("{today}", today)
        prompt = prompt.replace("{yesterday}", yesterday)
        return prompt
    
    # 1. Try to load from prompts folder (priority)
    prompts_path = os.path.join(os.path.dirname(__file__), "prompts", system_prompt_file)
    if os.path.exists(prompts_path):
        with open(prompts_path, "r", encoding="utf-8") as f:
            prompt = f.read().strip()
        return _process_prompt(prompt)
    
    # 2. Try to load from model directory
    prompt_path = os.path.join(model_path, system_prompt_file)
    if os.path.exists(prompt_path):
        with open(prompt_path, "r", encoding="utf-8") as f:
            prompt = f.read().strip()
        return _process_prompt(prompt)
    
    # 3. Fallback to base path from FileConfig
    if file_config and hasattr(file_config, 'BASE_PATH'):
        prompt_path = os.path.join(file_config.BASE_PATH, system_prompt_file)
        if os.path.exists(prompt_path):
            with open(prompt_path, "r", encoding="utf-8") as f:
                prompt = f.read().strip()
            return _process_prompt(prompt)
    
    return DEFAULT_SYSTEM_PROMPT


def set_chat_template_from_file(tokenizer, model_path, model_type="ministral_3_3b_instruct", debug=False):
    """
    Load chat_template.jinja from the model directory and assign it to the tokenizer.
    Uses FileConfig.CHAT_TEMPLATE for the filename.
    """
    if tokenizer is None:
        return False
    
    file_config = get_file_config(model_type)
    chat_template_file = file_config.CHAT_TEMPLATE if file_config and hasattr(file_config, 'CHAT_TEMPLATE') else "chat_template.jinja"
    
    template_path = os.path.join(model_path, chat_template_file)
    if not os.path.exists(template_path):
        # Fallback to base path
        if file_config and hasattr(file_config, 'BASE_PATH'):
            template_path = os.path.join(file_config.BASE_PATH, chat_template_file)
            if not os.path.exists(template_path):
                return False
        else:
            return False
    
    try:
        with open(template_path, "r", encoding="utf-8") as f:
            chat_template = f.read().strip()
        if not chat_template:
            return False
        tokenizer.chat_template = chat_template
        if debug:
            print(f"[DEBUG] Chat template loaded from {template_path}")
        return True
    except Exception as e:
        print(f"[WARNING] Failed to load chat template: {e}")
        return False


def build_messages(history, user_input, system_prompt=None):
    """
    Build the chat message list from history and current input.
    """
    messages = []
    
    # System prompt
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    
    # History
    for turn in history:
        messages.append({"role": turn["role"], "content": turn["content"]})
    
    # Current user input
    messages.append({"role": "user", "content": user_input})
    
    return messages


def build_prompt(messages, tokenizer=None):
    """
    Build prompt string in manual format (fallback path).
    """
    # Fallback to manual format (match chat_template.jinja)
    bos = "<s>"
    eos = "</s>"
    if tokenizer is not None:
        if tokenizer.bos_token:
            bos = tokenizer.bos_token
        if tokenizer.eos_token:
            eos = tokenizer.eos_token
    prompt = ""
    
    for msg in messages:
        role = msg["role"]
        content = msg["content"]
        prompt += f"{bos}{role}\n{content}{eos}\n"
    
    # Start assistant response
    prompt += f"{bos}assistant\n"
    
    return prompt


def retry_llm_generation(
    history,
    assistant_response,
    feedback_message,
    system_prompt,
    tokenizer,
    model,
    args,
    tools_schema=None,
    max_retries=None
):
    """
    Unified function to provide feedback to LLM and request regeneration.
    
    Args:
        history: Conversation history
        assistant_response: Previous assistant response (feedback target)
        feedback_message: Feedback message to send to LLM
        system_prompt: System prompt
        tokenizer: Tokenizer
        model: Model
        args: Generation arguments
        tools_schema: Tool schema (optional)
        max_retries: Maximum retry attempts (default: MAX_RETRY_ATTEMPTS)
    
    Yields:
        (retry_num, response) tuple - retry number and response for each attempt
    """
    if max_retries is None:
        max_retries = MAX_RETRY_ATTEMPTS
    
    retry_history = history.copy()
    retry_history.append({'role': 'assistant', 'content': assistant_response})
    retry_history.append({'role': 'user', 'content': feedback_message})
    
    for retry in range(max_retries):
        messages = build_messages(retry_history, "", system_prompt)
        inputs = build_inputs(messages, tokenizer, args, tools=tools_schema)
        
        response = ""
        for chunk in generate_with_refusal_streaming(model, tokenizer, inputs, args):
            if chunk.get('token'):
                response += chunk['token']
            if chunk.get('done'):
                break
        
        yield retry + 1, response
        
        # Update history for next retry
        retry_history.append({'role': 'assistant', 'content': response})
        retry_history.append({'role': 'user', 'content': feedback_message})


def build_inputs(messages, tokenizer, args, tools=None, inject_tool_call_prefix=False):
    """
    Build tokenized inputs using chat_template when available.
    Applies max_context truncation (older tokens removed first).
    
    Args:
        messages: Chat messages
        tokenizer: Tokenizer instance
        args: Arguments
        tools: Optional list of tool schemas for tool-enabled models
        inject_tool_call_prefix: If True, append [TOOL_CALLS] token at the end
    """
    inputs = None
    prompt_text = None

    if tokenizer is not None and hasattr(tokenizer, "apply_chat_template"):
        try:
            # Build template kwargs
            template_kwargs = {
                "tokenize": True,
                "add_generation_prompt": True,
                "return_tensors": "pt",
            }
            
            # Add tools if provided (for models that support tool calling)
            if tools:
                template_kwargs["tools"] = tools
            
            inputs = tokenizer.apply_chat_template(messages, **template_kwargs)
            
            # Inject [TOOL_CALLS] token at the end if requested
            if inject_tool_call_prefix and inputs is not None:
                # [TOOL_CALLS] token ID is 9
                tool_calls_token_id = 9
                tool_token = torch.tensor([[tool_calls_token_id]], dtype=torch.long)
                
                if isinstance(inputs, torch.Tensor):
                    # Direct tensor output
                    if inputs.dim() == 1:
                        inputs = inputs.unsqueeze(0)
                    inputs = torch.cat([inputs, tool_token], dim=1)
                elif isinstance(inputs, dict) and 'input_ids' in inputs:
                    # Dict with input_ids
                    input_ids = inputs['input_ids']
                    if input_ids.dim() == 1:
                        input_ids = input_ids.unsqueeze(0)
                    inputs['input_ids'] = torch.cat([input_ids, tool_token], dim=1)
                    if 'attention_mask' in inputs:
                        attention_token = torch.ones((1, 1), dtype=inputs['attention_mask'].dtype)
                        inputs['attention_mask'] = torch.cat([inputs['attention_mask'], attention_token], dim=1)
                elif hasattr(inputs, 'input_ids'):
                    # BatchEncoding or similar
                    input_ids = inputs.input_ids
                    if isinstance(input_ids, torch.Tensor):
                        if input_ids.dim() == 1:
                            input_ids = input_ids.unsqueeze(0)
                        inputs.input_ids = torch.cat([input_ids, tool_token], dim=1)
                        if hasattr(inputs, 'attention_mask') and inputs.attention_mask is not None:
                            attention_token = torch.ones((1, 1), dtype=inputs.attention_mask.dtype)
                            inputs.attention_mask = torch.cat([inputs.attention_mask, attention_token], dim=1)
            
            if args.debug_prompt:
                debug_kwargs = {
                    "tokenize": False,
                    "add_generation_prompt": True,
                }
                if tools:
                    debug_kwargs["tools"] = tools
                prompt_text = tokenizer.apply_chat_template(messages, **debug_kwargs)
        except Exception as e:
            print(f"[Warning] apply_chat_template failed: {e}")

    if inputs is None:
        prompt_text = build_prompt(messages, tokenizer)
        inputs = tokenizer(prompt_text, return_tensors="pt", add_special_tokens=False)

    if args.debug_prompt:
        print("\n" + "=" * 20 + " Prompt " + "=" * 20)
        print(prompt_text)
        print("=" * 20 + " Tokens " + "=" * 20)
        token_ids = inputs.input_ids[0][:args.debug_prompt_tokens].tolist()
        print(token_ids)
        print("=" * 50 + "\n")

    # Apply max_context truncation (remove older tokens from the beginning)
    max_context = getattr(args, 'max_context', 32768)
    if max_context > 0:
        # Handle both tensor and dict formats
        if isinstance(inputs, dict):
            input_ids = inputs['input_ids']
        elif hasattr(inputs, 'input_ids'):
            # Handle BatchEncoding or similar objects
            input_ids = inputs.input_ids
        else:
            input_ids = inputs
        
        # Get sequence length safely
        if hasattr(input_ids, 'dim'):
            seq_len = input_ids.shape[1] if input_ids.dim() > 1 else input_ids.shape[0]
        else:
            seq_len = input_ids.shape[1] if len(input_ids.shape) > 1 else input_ids.shape[0]
        
        if seq_len > max_context:
            truncate_amount = seq_len - max_context
            print(f"[Context] Truncating {truncate_amount} tokens from beginning (total: {seq_len} -> {max_context})")
            
            if isinstance(inputs, dict):
                inputs['input_ids'] = input_ids[:, truncate_amount:]
                if 'attention_mask' in inputs:
                    inputs['attention_mask'] = inputs['attention_mask'][:, truncate_amount:]
            elif hasattr(inputs, 'input_ids'):
                inputs.input_ids = input_ids[:, truncate_amount:]
                if hasattr(inputs, 'attention_mask') and inputs.attention_mask is not None:
                    inputs.attention_mask = inputs.attention_mask[:, truncate_amount:]
            else:
                if len(input_ids.shape) > 1:
                    inputs = input_ids[:, truncate_amount:]
                else:
                    inputs = input_ids[truncate_amount:]

    return inputs


def process_image_from_base64(base64_data):
    """
    Convert Base64 image data to PIL Image.
    
    Args:
        base64_data: Base64 encoded image string (data:image/...;base64,...)
    
    Returns:
        PIL Image object
    """
    if not PIL_AVAILABLE:
        raise ImportError("PIL is required for image processing")
    
    # Remove data URL prefix if present
    if ',' in base64_data:
        base64_data = base64_data.split(',')[1]
    
    image_bytes = base64.b64decode(base64_data)
    image = Image.open(io.BytesIO(image_bytes))
    
    # Convert to RGB if necessary
    if image.mode != 'RGB':
        image = image.convert('RGB')
    
    return image


def build_multimodal_messages(history, user_text, images, system_prompt=None):
    """
    Build chat messages with image content for multimodal models.
    
    Args:
        history: Previous conversation turns
        user_text: Current user text input
        images: List of PIL Image objects
        system_prompt: Optional system prompt
    
    Returns:
        List of messages in multimodal format
    """
    messages = []
    
    # System prompt
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    
    # History (text only for now)
    for turn in history:
        messages.append({"role": turn["role"], "content": turn["content"]})
    
    # Current user input with images
    if images:
        # Multimodal content format
        content = []
        for img in images:
            content.append({"type": "image", "image": img})
        if user_text:
            content.append({"type": "text", "text": user_text})
        messages.append({"role": "user", "content": content})
    else:
        messages.append({"role": "user", "content": user_text})
    
    return messages


def build_inputs_multimodal(messages, images, processor, tokenizer, args, device='cuda'):
    """
    Build inputs for multimodal model using processor.
    Applies max_context truncation (older tokens removed first).
    
    Args:
        messages: Chat messages (may contain image content)
        images: List of PIL Image objects
        processor: PixtralProcessor or similar
        tokenizer: Tokenizer (fallback)
        args: Arguments
        device: Target device
    
    Returns:
        Dictionary with input_ids, attention_mask, and optionally pixel_values
    """
    if images and processor is not None:
        try:
            # Build prompt text for the processor
            prompt_text = processor.apply_chat_template(
                messages,
                tokenize=False,
                add_generation_prompt=True,
            )
            
            # Process text and images together
            inputs = processor(
                text=prompt_text,
                images=images,
                return_tensors="pt",
                padding=True,
            )
            
            # Move to device
            inputs = {k: v.to(device) if hasattr(v, 'to') else v for k, v in inputs.items()}
            
            # Apply max_context truncation (remove older tokens from the beginning)
            # Note: For multimodal, truncating may affect image tokens - use with caution
            max_context = getattr(args, 'max_context', 32768)
            if max_context > 0 and 'input_ids' in inputs:
                seq_len = inputs['input_ids'].shape[1]
                if seq_len > max_context:
                    truncate_amount = seq_len - max_context
                    print(f"[Context] Truncating {truncate_amount} tokens from beginning (total: {seq_len} -> {max_context})")
                    print(f"[Warning] Multimodal truncation may affect image tokens")
                    inputs['input_ids'] = inputs['input_ids'][:, truncate_amount:]
                    if 'attention_mask' in inputs:
                        inputs['attention_mask'] = inputs['attention_mask'][:, truncate_amount:]
            
            if args.debug_prompt:
                print("\n" + "=" * 20 + " Multimodal Prompt " + "=" * 20)
                print(prompt_text[:500] + "..." if len(prompt_text) > 500 else prompt_text)
                print(f"Images: {len(images)}")
                print("=" * 50 + "\n")
            
            return inputs
        except Exception as e:
            print(f"[Warning] Multimodal processing failed: {e}")
            # Fall back to text-only
    
    # Text-only fallback
    return build_inputs(messages, tokenizer, args)


def generate_response(model, tokenizer, inputs, args, streaming=True):
    """
    Generate response from the model.
    
    Args:
        model: The model to generate from
        tokenizer: Tokenizer for decoding
        inputs: Tokenized inputs
        args: Generation arguments
        streaming: If True, print tokens as they are generated
    
    Returns:
        The generated response text
    """
    # Check if refusal mechanism is enabled (default: enabled)
    use_refusal = not getattr(args, 'no_refusal', False)
    
    if use_refusal:
        # Use token-by-token generation with refusal mechanism
        return generate_with_refusal(model, tokenizer, inputs, args)
    
    # Standard generation path (refusal disabled)
    device = get_model_device(model)
    inputs = inputs.to(device)
    input_length = inputs.input_ids.shape[1]
    
    # Create streamer for real-time output
    # Note: skip_special_tokens=False to preserve tool tokens like [TOOL_CALLS], [ARGS]
    # Meta tokens (<s>, </s>, etc.) may appear but won't affect functionality
    streamer = None
    if streaming:
        streamer = TextStreamer(
            tokenizer,
            skip_prompt=True,  # Don't print the input prompt
            skip_special_tokens=False  # Preserve tool tokens
        )
    
    with torch.no_grad():
        outputs = model.generate(
            **inputs,
            max_new_tokens=args.max_length,
            temperature=args.temperature,
            top_k=args.top_k,
            top_p=0.9,
            do_sample=True if args.temperature > 0 else False,
            pad_token_id=tokenizer.eos_token_id,
            eos_token_id=tokenizer.eos_token_id,
            use_cache=False,
            streamer=streamer,
        )
    
    # Extract only the new tokens (response) for history
    # Use skip_special_tokens=False to preserve tool tokens, then strip meta tokens
    response = tokenizer.decode(outputs[0][input_length:], skip_special_tokens=False)
    response = strip_meta_tokens(response)
    return response.strip()


def run_vision_inference(image_path: str, prompt: str = "Describe this image in detail.") -> str:
    """Run multimodal inference on a single image. Returns text result.
    
    Called by the view_image tool. Uses the globally loaded model and processor.
    
    Args:
        image_path: Path to image file (absolute, or relative like /uploads/filename)
        prompt: Text prompt for image analysis
    
    Returns:
        Generated text description
        
    Raises:
        FileNotFoundError: If image file doesn't exist
        RuntimeError: If vision encoder is not available
    """
    global global_model, global_processor, global_tokenizer, global_args, model_lock

    if global_processor is None:
        raise RuntimeError("Vision encoder not available. Text-only model loaded.")
    if global_model is None:
        raise RuntimeError("Model not loaded.")

    # Resolve path: /uploads/filename -> UPLOADS_DIR/filename
    if image_path.startswith('/uploads/'):
        resolved = os.path.join(UPLOADS_DIR, image_path[len('/uploads/'):])
    else:
        resolved = image_path

    if not os.path.isfile(resolved):
        raise FileNotFoundError(f"Image file not found: {image_path}")

    img = Image.open(resolved).convert('RGB')
    images = [img]

    messages = build_multimodal_messages([], prompt, images)
    inputs = build_inputs_multimodal(
        messages, images, global_processor, global_tokenizer, global_args
    )

    with model_lock:
        response = generate_response(global_model, global_tokenizer, inputs, global_args, streaming=False)

    return response


def print_history(history):
    """
    Print conversation history.
    """
    if not history:
        print("\n[No conversation history yet]\n")
        return
    
    print("\n" + "="*50)
    print("Conversation History")
    print("="*50)
    for i, turn in enumerate(history):
        role = turn["role"].capitalize()
        content = turn["content"]
        print(f"\n[{i+1}] {role}: {content}")
    print("\n" + "="*50 + "\n")


def get_model_name(model_path):
    """
    Extract model name from folder path.
    """
    # Get the last folder name from path
    return os.path.basename(os.path.normpath(model_path))


def save_conversation(model_name, history, log_dir="inference_log"):
    """
    Save conversation history to JSON file.
    """
    if not history:
        print("No conversation to save.")
        return None
    
    # Create log directory if not exists
    if not os.path.exists(log_dir):
        os.makedirs(log_dir, exist_ok=True)
    
    # Generate filename
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%H%M%S")
    filename = f"{model_name}-{date_str}-{time_str}.json"
    filepath = os.path.join(log_dir, filename)
    
    # Prepare data
    data = {
        "model": model_name,
        "timestamp": now.isoformat(),
        "conversation": history
    }
    
    # Save to file
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"Conversation saved to: {filepath}")
    return filepath


def run_chat(model, tokenizer, model_name, system_prompt, args):
    """
    Main chat loop with Agent/Plan mode support.
    """
    cli_mode = 'agent'
    
    print("\n" + "="*50)
    print(f"Chat with {model_name}")
    print("="*50)
    print("Commands:")
    print("  /agent   - Switch to Agent mode (direct chat)")
    print("  /plan    - Switch to Plan mode (tool routing)")
    print("  /clear   - Clear conversation history")
    print("  /history - Show conversation history")
    print("  /save    - Save conversation now")
    print("  /exit    - Save and quit")
    print("="*50)
    print(f"[Current mode: Agent]\n")
    
    conversation_history = []
    
    while True:
        try:
            prompt_prefix = "[Agent]" if cli_mode == 'agent' else "[Plan]"
            user_input = input(f"{prompt_prefix} You: ").strip()
        except (KeyboardInterrupt, EOFError):
            print("\n")
            save_conversation(model_name, conversation_history)
            print("Goodbye!")
            break
        
        if not user_input:
            continue
        
        # Special commands
        if user_input == '/exit':
            save_conversation(model_name, conversation_history)
            print("Goodbye!")
            break
        
        if user_input == '/clear':
            conversation_history = []
            print("[History cleared]\n")
            continue
        
        if user_input == '/history':
            print_history(conversation_history)
            continue
        
        if user_input == '/save':
            save_conversation(model_name, conversation_history)
            continue
        
        if user_input == '/agent':
            cli_mode = 'agent'
            print("[Switched to Agent mode]\n")
            continue
        
        if user_input == '/plan':
            cli_mode = 'plan'
            print("[Switched to Plan mode]\n")
            continue
        
        # Determine tools based on mode
        tools_schema = None
        current_sys_prompt = system_prompt
        
        if cli_mode == 'plan':
            plan_prompt_path = os.path.join(os.path.dirname(__file__), "prompts", "PLAN_SYSTEM_PROMPT.txt")
            if os.path.exists(plan_prompt_path):
                with open(plan_prompt_path, "r", encoding="utf-8") as f:
                    current_sys_prompt = f.read().strip()
            if TOOLS_AVAILABLE:
                from tools.base import get_plan_schema
                tools_schema = get_plan_schema()
        
        # Build tokenized inputs
        messages = build_messages(conversation_history, user_input, current_sys_prompt)
        inputs = build_inputs(messages, tokenizer, args, tools=tools_schema)
        
        # Generate response with streaming
        try:
            print("\nAssistant: ", end="", flush=True)
            response = generate_response(model, tokenizer, inputs, args, streaming=True)
            print()
            
            conversation_history.append({"role": "user", "content": user_input})
            conversation_history.append({"role": "assistant", "content": response})
            
            # In plan mode, check for tool calls in response
            if cli_mode == 'plan' and tools_schema and '[TOOL_CALLS]' in response:
                try:
                    tool_call_text = response.split('[TOOL_CALLS]', 1)[1].strip()
                    import re
                    tool_calls = re.findall(r'\{[^{}]+\}', tool_call_text)
                    for tc_str in tool_calls:
                        tc = json.loads(tc_str)
                        tool_name = tc.get('name', '')
                        tool_args = tc.get('arguments', {})
                        print(f"\n[Tool Call: {tool_name}]")
                        if tool_name == 'create_plan':
                            goal = tool_args.get('goal', '')
                            steps = tool_args.get('steps', [])
                            print(f"Goal: {goal}")
                            for i, step in enumerate(steps, 1):
                                print(f"  {i}. {step.get('name', '')} [{step.get('tool', '')}]")
                            print()
                except Exception as e:
                    print(f"\n[Tool parse error: {e}]")
            
        except Exception as e:
            print(f"\n[Error generating response: {e}]\n")
            traceback.print_exc()


# ============================================
# Web UI Mode
# ============================================

# Global state for web server
global_model = None
global_tokenizer = None
global_processor = None  # For multimodal input processing
global_args = None
global_system_prompt = None
global_model_name = None
global_stop_generation = False
model_lock = threading.RLock()

INFERENCE_UI_DIR = os.path.join(os.path.dirname(__file__), "inference_ui")
LOGS_DIR = os.path.join(INFERENCE_UI_DIR, "logs")
UPLOADS_DIR = os.path.join(INFERENCE_UI_DIR, "uploads")
OUTPUTS_DIR = os.path.join(INFERENCE_UI_DIR, "outputs")
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(OUTPUTS_DIR, exist_ok=True)


def _execute_code_subprocess(code, language, conv_id, step_index):
    """Execute Python/R code via subprocess and capture stdout, stderr, figures, tables.

    This is a pure function with no HTTP or model_lock dependencies.
    The caller is responsible for any locking if needed.
    """
    out_dir = os.path.join(OUTPUTS_DIR, str(conv_id), f'step_{step_index}')
    out_dir = out_dir.replace('\\', '/')
    os.makedirs(out_dir, exist_ok=True)

    if language == 'r':
        ext = '.R'
        cmd_prefix = ['Rscript']
        preamble = ''
        postamble = ''
    else:
        ext = '.py'
        cmd_prefix = [sys.executable]
        preamble = (
            "import os as _os\n"
            "import json as _json\n"
            "import matplotlib as _mpl\n"
            "_mpl.use('Agg')\n"
            "import matplotlib.pyplot as _plt\n"
            f"_out_dir = {repr(out_dir)}\n"
            f"_data_dir = {repr(out_dir)}\n"
            "_fig_count = [0]\n"
            "# Load previous step results if available\n"
            "results = {}\n"
            "_prev_path = _os.path.join(_out_dir, '_prev_data.json')\n"
            "if _os.path.isfile(_prev_path):\n"
            "    try:\n"
            "        with open(_prev_path, 'r', encoding='utf-8') as _pf:\n"
            "            _prev_list = _json.load(_pf)\n"
            "        for _item in _prev_list:\n"
            "            _sn = _item.get('step_num', _item.get('step', 0))\n"
            "            results[_sn] = _item.get('result', {}).get('result', _item.get('result', {}))\n"
            "    except Exception:\n"
            "        pass\n"
            "# Patch plt.show to auto-save figures\n"
            "_orig_show = _plt.show\n"
            "def _patched_show(*a, **kw):\n"
            "    _fig_count[0] += 1\n"
            "    _plt.savefig(f'{_out_dir}/fig_{_fig_count[0]}.png', dpi=100, bbox_inches='tight')\n"
            "    _plt.close('all')\n"
            "_plt.show = _patched_show\n"
            "# Patch plt.savefig to also copy to _out_dir\n"
            "_orig_savefig = _plt.savefig\n"
            "def _patched_savefig(fname, *a, **kw):\n"
            "    _orig_savefig(fname, *a, **kw)\n"
            "    if isinstance(fname, str) and not _os.path.isabs(fname):\n"
            "        _fig_count[0] += 1\n"
            "        dest = f'{_out_dir}/fig_{_fig_count[0]}.png'\n"
            "        if _os.path.abspath(fname) != _os.path.abspath(dest):\n"
            "            try:\n"
            "                import shutil as _shutil\n"
            "                _shutil.copy2(_os.path.abspath(fname), dest)\n"
            "            except Exception:\n"
            "                pass\n"
            "_plt.savefig = _patched_savefig\n"
        )
        postamble = (
            "\n# --- auto-save cleanup ---\n"
            "import matplotlib.pyplot as _plt2\n"
            "if _plt2.get_fignums():\n"
            "    _fig_count[0] += 1\n"
            "    _plt2.savefig(f'{_out_dir}/fig_{_fig_count[0]}.png', dpi=100, bbox_inches='tight')\n"
            "    _plt2.close('all')\n"
            "try:\n"
            "    import pandas as _pd\n"
            "    _tbl_count = 0\n"
            "    for _vname, _vval in list(locals().items()):\n"
            "        if isinstance(_vval, _pd.DataFrame) and not _vname.startswith('_'):\n"
            "            _tbl_count += 1\n"
            "            _vval.to_csv(f'{_out_dir}/table_{_tbl_count}.csv', index=False)\n"
            "except ImportError:\n"
            "    pass\n"
        )

    if language == 'python':
        has_main_def = bool(re.search(r'^def\s+main\s*\(', code, re.MULTILINE))
        has_main_call = bool(re.search(r'(?:^|\n)(?:if\s+__name__.*)?main\s*\(', code))
        if has_main_def and not has_main_call:
            code += '\n\nmain()\n'

    full_code = preamble + code + postamble
    tmp_file = None
    try:
        tmp_file = tempfile.NamedTemporaryFile(mode='w', suffix=ext, delete=False, encoding='utf-8')
        tmp_file.write(full_code)
        tmp_file.close()

        proc = subprocess.run(
            cmd_prefix + [tmp_file.name],
            capture_output=True, text=True, timeout=60,
            cwd=out_dir
        )

        stdout = proc.stdout or ''
        stderr = proc.stderr or ''
        success = proc.returncode == 0
        print(f"[exec] returncode={proc.returncode}, stdout={len(stdout)} chars, stderr={len(stderr)} chars")
        print(f"[exec] cwd={out_dir}")
        if stderr.strip():
            print(f"[exec] stderr preview: {stderr[:300]}")
    except subprocess.TimeoutExpired:
        stdout = ''
        stderr = 'Code execution timed out (60s limit)'
        success = False
    except Exception as e:
        stdout = ''
        stderr = str(e)
        success = False
    finally:
        if tmp_file and os.path.exists(tmp_file.name):
            os.unlink(tmp_file.name)

    figures = []
    tables = []
    if os.path.isdir(out_dir):
        all_files = sorted(os.listdir(out_dir))
        for f in all_files:
            if f.startswith('_'):
                continue
            url = f'/api/outputs/{conv_id}/step_{step_index}/{f}'
            if f.endswith('.png'):
                figures.append(url)
            elif f.endswith('.csv'):
                tables.append(url)
        print(f"[exec] output files: {[f for f in all_files if not f.startswith('_')]}, figures={len(figures)}, tables={len(tables)}")

    return {
        'success': success,
        'stdout': stdout,
        'stderr': stderr,
        'figures': figures,
        'tables': tables
    }


def _extract_text_from_file(filepath, filename):
    """Extract readable text from uploaded files with graceful library fallback."""
    ext = os.path.splitext(filename)[1].lower()

    if ext in ('.csv', '.xml', '.json', '.txt'):
        try:
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                return f.read(), 'raw'
        except Exception:
            return f'[Failed to read {filename}]', 'error'

    if ext == '.pdf':
        try:
            import fitz
            doc = fitz.open(filepath)
            pages = [page.get_text() for page in doc]
            doc.close()
            return '\n\n--- Page Break ---\n\n'.join(pages), 'pymupdf'
        except ImportError:
            return f'[PDF file: {filename} - install PyMuPDF for text extraction]', 'unavailable'
        except Exception as e:
            return f'[PDF read error: {e}]', 'error'

    if ext in ('.doc', '.docx'):
        try:
            from docx import Document
            doc = Document(filepath)
            text = '\n'.join(p.text for p in doc.paragraphs if p.text)
            return text, 'python-docx'
        except ImportError:
            return f'[Word file: {filename} - install python-docx for text extraction]', 'unavailable'
        except Exception as e:
            return f'[Word read error: {e}]', 'error'

    if ext in ('.xlsx', '.xls'):
        try:
            from openpyxl import load_workbook
            wb_f = load_workbook(filepath, data_only=False)
            wb_v = load_workbook(filepath, data_only=True)
            sheets_text = []
            for sname in wb_f.sheetnames:
                ws_f = wb_f[sname]
                ws_v = wb_v[sname]
                if ws_f.max_row is None or ws_f.max_column is None:
                    continue
                rows = []
                for r in range(1, ws_f.max_row + 1):
                    row = []
                    for c in range(1, ws_f.max_column + 1):
                        fval = ws_f.cell(r, c).value
                        cval = ws_v.cell(r, c).value
                        if isinstance(fval, str) and fval.startswith('='):
                            row.append(f'{cval} [{fval}]' if cval is not None else f'[{fval}]')
                        else:
                            row.append(str(fval) if fval is not None else '')
                    rows.append(row)
                if rows:
                    header = '| ' + ' | '.join(rows[0]) + ' |'
                    sep = '| ' + ' | '.join('---' for _ in rows[0]) + ' |'
                    body_lines = ['| ' + ' | '.join(r) + ' |' for r in rows[1:]]
                    sheets_text.append(f'## {sname}\n{header}\n{sep}\n' + '\n'.join(body_lines))
            wb_f.close()
            wb_v.close()
            return '\n\n'.join(sheets_text), 'openpyxl'
        except ImportError:
            return f'[Excel file: {filename} - install openpyxl for text extraction]', 'unavailable'
        except Exception as e:
            return f'[Excel read error: {e}]', 'error'

    return f'[Unsupported format: {ext}]', 'unsupported'


class ConversationManager:
    """Manage conversation logs."""
    
    def __init__(self, logs_dir):
        self.logs_dir = logs_dir
        os.makedirs(logs_dir, exist_ok=True)
        self._list_cache = None
        self._list_cache_mtime = 0
    
    def _invalidate_cache(self):
        self._list_cache = None
    
    def list_conversations(self):
        """List all conversations (cached, re-reads only when directory changes)."""
        try:
            dir_mtime = os.path.getmtime(self.logs_dir)
        except OSError:
            dir_mtime = 0
        
        if self._list_cache is not None and dir_mtime <= self._list_cache_mtime:
            return self._list_cache
        
        conversations = []
        for filename in os.listdir(self.logs_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(self.logs_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        conversations.append({
                            'id': data.get('id', filename[:-5]),
                            'title': data.get('title', 'New Chat'),
                            'created_at': data.get('created_at'),
                            'updated_at': data.get('updated_at')
                        })
                except:
                    pass
        
        # Sort by updated_at descending
        conversations.sort(key=lambda x: x.get('updated_at', ''), reverse=True)
        self._list_cache = conversations
        self._list_cache_mtime = dir_mtime
        return conversations
    
    def get_conversation(self, conv_id):
        """Get a specific conversation."""
        filepath = os.path.join(self.logs_dir, f"{conv_id}.json")
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        return None
    
    def create_conversation(self):
        """Create a new conversation."""
        conv_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        data = {
            'id': conv_id,
            'title': 'New Chat',
            'model': global_model_name or 'unknown',
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'messages': []
        }
        self.save_conversation(conv_id, data)
        return data
    
    def save_conversation(self, conv_id, data):
        """Save a conversation."""
        data['updated_at'] = datetime.now().isoformat()
        filepath = os.path.join(self.logs_dir, f"{conv_id}.json")
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self._invalidate_cache()
    
    def delete_conversation(self, conv_id):
        """Delete a conversation."""
        filepath = os.path.join(self.logs_dir, f"{conv_id}.json")
        if os.path.exists(filepath):
            os.remove(filepath)
            self._invalidate_cache()
            return True
        return False
    
    def add_message(self, conv_id, role, content, files=None):
        """Add a message to a conversation."""
        data = self.get_conversation(conv_id)
        if data:
            msg = {
                'role': role,
                'content': content
            }
            if files:
                msg['files'] = files
            data['messages'].append(msg)
            # Update title from first user message
            if role == 'user' and data['title'] == 'New Chat':
                # Extract text part for title (exclude file references)
                title_text = content
                # Remove [Image: xxx] and [Audio: xxx] references
                title_text = re.sub(r'\[Image: [^\]]+\]\s*', '', title_text)
                title_text = re.sub(r'\[Audio: [^\]]+\]\s*', '', title_text)
                title_text = title_text.strip()
                
                if title_text:
                    data['title'] = title_text[:50] + ('...' if len(title_text) > 50 else '')
                else:
                    # Only images/audio, no text - use default title
                    data['title'] = 'Image Chat'
            self.save_conversation(conv_id, data)
            return True
        return False
    
    def rename_conversation(self, conv_id, new_title):
        """Rename a conversation."""
        data = self.get_conversation(conv_id)
        if data:
            data['title'] = new_title
            self.save_conversation(conv_id, data)
            return True
        return False
    
    def truncate_messages(self, conv_id, from_index):
        """Delete messages from a specific index onwards."""
        data = self.get_conversation(conv_id)
        if data and 0 <= from_index < len(data.get('messages', [])):
            data['messages'] = data['messages'][:from_index]
            self.save_conversation(conv_id, data)
            return True
        return False
    
    def replace_last_plan_message(self, conv_id, new_content):
        """Replace the last [TOOL_CALLS]...create_plan message with new content.
        Used to replace the initial create_plan message with [PLAN_COMPLETE] data,
        preventing duplicate plan boxes in the chat."""
        data = self.get_conversation(conv_id)
        if not data:
            return False
        messages = data.get('messages', [])
        for i in range(len(messages) - 1, -1, -1):
            msg = messages[i]
            content = msg.get('content') or ''
            if msg.get('role') == 'assistant' and '[TOOL_CALLS]' in content and 'create_plan' in content:
                msg['content'] = new_content
                self.save_conversation(conv_id, data)
                return True
        return False

    def clear_conversation(self, conv_id):
        """Clear messages in a conversation."""
        data = self.get_conversation(conv_id)
        if data:
            data['messages'] = []
            data['title'] = 'New Chat'
            self.save_conversation(conv_id, data)
            return True
        return False

    def update_plan_analysis(self, conv_id, analysis_text):
        """Append analysis text to the last PLAN_COMPLETE message in a conversation."""
        data = self.get_conversation(conv_id)
        if not data:
            return False
        messages = data.get('messages', [])
        for i in range(len(messages) - 1, -1, -1):
            msg = messages[i]
            if msg.get('role') == 'assistant' and '[PLAN_COMPLETE]' in (msg.get('content') or ''):
                content = msg['content']
                match_idx = content.find('[PLAN_COMPLETE]')
                if match_idx == -1:
                    continue
                json_str = content[match_idx + len('[PLAN_COMPLETE]'):].strip()
                try:
                    plan_json = json.loads(json_str)
                    plan_json['analysis'] = analysis_text
                    prefix = content[:match_idx]
                    msg['content'] = prefix + '[PLAN_COMPLETE]' + json.dumps(plan_json, ensure_ascii=False)
                    self.save_conversation(conv_id, data)
                    return True
                except json.JSONDecodeError:
                    return False
        return False


conversation_manager = None


class InferenceChatHandler(BaseHTTPRequestHandler):
    """HTTP handler for inference chat web UI."""
    
    def log_message(self, format, *args):
        pass  # Suppress default logging
    
    def send_json(self, data, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode('utf-8'))
    
    def send_error_json(self, message, status=400):
        self.send_json({'error': message}, status)
    
    def serve_static_file(self, filepath, content_type=None):
        if not os.path.exists(filepath):
            self.send_error(404, 'File not found')
            return
        
        if content_type is None:
            ext = os.path.splitext(filepath)[1].lower()
            content_types = {
                '.html': 'text/html',
                '.css': 'text/css',
                '.js': 'application/javascript',
                '.json': 'application/json',
                '.png': 'image/png',
                '.ico': 'image/x-icon'
            }
            content_type = content_types.get(ext, 'application/octet-stream')
        
        self.send_response(200)
        self.send_header('Content-Type', content_type)
        self.end_headers()
        with open(filepath, 'rb') as f:
            self.wfile.write(f.read())
    
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        
        # Static files
        if path == '/' or path == '/index.html':
            self.serve_static_file(os.path.join(INFERENCE_UI_DIR, 'index.html'))
        elif path == '/style.css':
            self.serve_static_file(os.path.join(INFERENCE_UI_DIR, 'style.css'))
        elif path == '/app.js':
            self.serve_static_file(os.path.join(INFERENCE_UI_DIR, 'app.js'))
        elif path == '/graph.html':
            self.serve_static_file(os.path.join(INFERENCE_UI_DIR, 'graph.html'))
        elif path == '/node-graph.js':
            self.serve_static_file(os.path.join(INFERENCE_UI_DIR, 'node-graph.js'))
        elif path == '/node-graph.css':
            self.serve_static_file(os.path.join(INFERENCE_UI_DIR, 'node-graph.css'))
        elif path == '/i18n.js':
            self.serve_static_file(os.path.join(INFERENCE_UI_DIR, 'i18n.js'))
        elif path.startswith('/language/') and path.endswith('.json'):
            safe_name = os.path.basename(path)
            self.serve_static_file(os.path.join(INFERENCE_UI_DIR, 'language', safe_name))
        elif path.startswith('/nodes/') and path.endswith('.js'):
            safe_path = os.path.normpath(path.lstrip('/'))
            full_path = os.path.join(INFERENCE_UI_DIR, safe_path)
            if os.path.commonpath([INFERENCE_UI_DIR, full_path]) == INFERENCE_UI_DIR:
                self.serve_static_file(full_path, 'application/javascript')
            else:
                self.send_error(403, 'Forbidden')
        
        elif path.startswith('/lib/') and (path.endswith('.js') or path.endswith('.css')):
            safe_path = os.path.normpath(path.lstrip('/'))
            full_path = os.path.join(INFERENCE_UI_DIR, safe_path)
            if os.path.commonpath([INFERENCE_UI_DIR, full_path]) == INFERENCE_UI_DIR:
                self.serve_static_file(full_path)
            else:
                self.send_error(403, 'Forbidden')
        
        # API endpoints
        elif path == '/api/node-manifest':
            nodes_dir = os.path.join(INFERENCE_UI_DIR, 'nodes')
            helpers = []
            nodes = []
            for root, _dirs, files in os.walk(nodes_dir):
                for f in files:
                    if not f.endswith('.js'):
                        continue
                    if f == 'node-registry.js':
                        continue
                    rel = os.path.relpath(os.path.join(root, f), INFERENCE_UI_DIR).replace('\\', '/')
                    if f.endswith('-helper.js') or f.endswith('-util.js'):
                        helpers.append(rel)
                    else:
                        nodes.append(rel)
            helpers.sort()
            nodes.sort()
            self.send_json({'files': helpers + nodes})
        
        elif path == '/api/model':
            from utils.api_chat import get_active_model
            active = get_active_model()
            self.send_json({
                'model': global_model_name or active.get('local_model', ''),
                'mode': active.get('mode', 'local'),
                'provider': active.get('api_provider', ''),
                'api_model': active.get('api_model', '')
            })
        
        elif path == '/api/models':
            from utils.api_chat import get_active_model, get_provider_models
            active = get_active_model()
            local_models = list_available_models()
            api_models = get_provider_models()
            self.send_json({
                'local': local_models,
                'api': api_models,
                'active': {
                    'mode': active.get('mode', 'local'),
                    'local_model': active.get('local_model', global_model_name or ''),
                    'api_provider': active.get('api_provider', ''),
                    'api_model': active.get('api_model', '')
                }
            })
        
        elif path == '/api/api-keys':
            from utils.api_chat import get_api_keys_status
            self.send_json(get_api_keys_status())
        
        elif path == '/api/conversations':
            conversations = conversation_manager.list_conversations()
            self.send_json(conversations)
        
        elif path.startswith('/api/conversation/'):
            conv_id = path.split('/')[-1]
            data = conversation_manager.get_conversation(conv_id)
            if data:
                self.send_json(data)
            else:
                self.send_error_json('Conversation not found', 404)
        
        elif path == '/api/system_prompt':
            self.send_json({'system_prompt': global_system_prompt or ''})
        
        elif path == '/api/system_prompt/default':
            default_prompt = DEFAULT_SYSTEM_PROMPT
            file_config = get_file_config(global_args.model_type) if global_args else None
            sp_file = file_config.SYSTEM_PROMPT if file_config and hasattr(file_config, 'SYSTEM_PROMPT') else "SYSTEM_PROMPT.txt"
            if file_config and hasattr(file_config, 'BASE_PATH'):
                sp_path = os.path.join(file_config.BASE_PATH, sp_file)
                if os.path.exists(sp_path):
                    with open(sp_path, "r", encoding="utf-8") as f:
                        default_prompt = f.read().strip()
            model_path = getattr(global_args, 'model_path', None)
            if model_path:
                sp_path = os.path.join(model_path, sp_file)
                if os.path.exists(sp_path):
                    with open(sp_path, "r", encoding="utf-8") as f:
                        default_prompt = f.read().strip()
            self.send_json({'system_prompt': default_prompt})
        
        elif path == '/api/settings':
            self.send_json({
                'temperature': global_args.temperature,
                'max_length': global_args.max_length,
                'top_k': global_args.top_k,
                'max_context': global_args.max_context
            })
        
        elif path == '/api/data/list':
            file_list = []
            if os.path.isdir(UPLOADS_DIR):
                for fname in os.listdir(UPLOADS_DIR):
                    if fname.endswith('.extracted.txt'):
                        continue
                    fpath = os.path.join(UPLOADS_DIR, fname)
                    if os.path.isfile(fpath):
                        file_list.append({
                            'name': fname,
                            'size': os.path.getsize(fpath),
                            'mtime': os.path.getmtime(fpath)
                        })
            file_list.sort(key=lambda x: x['mtime'], reverse=True)
            self.send_json(file_list)
        
        elif path.startswith('/uploads/'):
            fname = os.path.basename(path)
            fpath = os.path.join(UPLOADS_DIR, fname)
            if os.path.commonpath([UPLOADS_DIR, os.path.abspath(fpath)]) == UPLOADS_DIR and os.path.isfile(fpath):
                self.serve_static_file(fpath)
            else:
                self.send_error(404, 'File not found')
        
        elif path.startswith('/api/outputs/'):
            rel = path[len('/api/outputs/'):]
            parts = rel.strip('/').split('/')
            if len(parts) == 3:
                full = os.path.normpath(os.path.join(OUTPUTS_DIR, *parts))
                if os.path.commonpath([OUTPUTS_DIR, full]) == OUTPUTS_DIR and os.path.isfile(full):
                    ct = 'image/png' if full.endswith('.png') else 'text/csv' if full.endswith('.csv') else 'application/octet-stream'
                    self.serve_static_file(full, ct)
                else:
                    self.send_error(404, 'File not found')
            elif len(parts) == 2:
                step_dir = os.path.normpath(os.path.join(OUTPUTS_DIR, parts[0], parts[1]))
                if os.path.commonpath([OUTPUTS_DIR, step_dir]) == OUTPUTS_DIR and os.path.isdir(step_dir):
                    figs = sorted([f'/api/outputs/{parts[0]}/{parts[1]}/{f}' for f in os.listdir(step_dir) if f.endswith('.png')])
                    tbls = sorted([f'/api/outputs/{parts[0]}/{parts[1]}/{f}' for f in os.listdir(step_dir) if f.endswith('.csv')])
                    self.send_json({'figures': figs, 'tables': tbls})
                else:
                    self.send_json({'figures': [], 'tables': []})
            else:
                self.send_error(404, 'Not found')
        
        else:
            self.send_error(404, 'Not found')
    
    def do_POST(self):
        global global_system_prompt, global_args, global_stop_generation
        parsed = urlparse(self.path)
        path = parsed.path
        
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length).decode('utf-8') if content_length > 0 else '{}'
        
        try:
            data = json.loads(body) if body else {}
        except json.JSONDecodeError:
            data = {}
        
        if path == '/api/new':
            conv = conversation_manager.create_conversation()
            self.send_json({'id': conv['id']})
        
        elif path == '/api/chat':
            self.handle_chat(data)
        
        elif path.startswith('/api/conversation/') and path.endswith('/clear'):
            conv_id = path.split('/')[-2]
            if conversation_manager.clear_conversation(conv_id):
                self.send_json({'success': True})
            else:
                self.send_error_json('Conversation not found', 404)
        
        elif path.startswith('/api/conversation/') and path.endswith('/rename'):
            conv_id = path.split('/')[-2]
            new_title = data.get('title', '').strip()
            if new_title and conversation_manager.rename_conversation(conv_id, new_title):
                self.send_json({'success': True})
            else:
                self.send_error_json('Failed to rename', 400)
        
        elif path.startswith('/api/conversation/') and path.endswith('/truncate'):
            conv_id = path.split('/')[-2]
            from_index = data.get('from_index')
            if from_index is not None and conversation_manager.truncate_messages(conv_id, int(from_index)):
                self.send_json({'success': True})
            else:
                self.send_error_json('Failed to truncate', 400)
        
        elif path == '/api/system_prompt':
            is_reset = data.get('reset', False)
            prompts_dir = os.path.join(os.path.dirname(__file__), "prompts")
            sp_path = os.path.join(prompts_dir, "SYSTEM_PROMPT.txt")
            
            if is_reset:
                # Delete custom file so model default is used on next restart
                try:
                    if os.path.exists(sp_path):
                        os.remove(sp_path)
                        print(f"[system_prompt] Deleted custom prompt file, reverting to model default")
                except Exception as e:
                    print(f"[Warning] Failed to delete custom prompt file: {e}")
                # Reload default from model directory
                global_system_prompt = load_system_prompt(
                    getattr(global_args, 'model_path', ''),
                    getattr(global_args, 'model_type', 'ministral_3_3b_instruct')
                )
            else:
                new_prompt = data.get('system_prompt', '')
                global_system_prompt = new_prompt
                os.makedirs(prompts_dir, exist_ok=True)
                try:
                    with open(sp_path, "w", encoding="utf-8") as f:
                        f.write(new_prompt)
                    print(f"[system_prompt] Saved custom prompt ({len(new_prompt)} chars)")
                except Exception as e:
                    print(f"[Warning] Failed to persist system prompt: {e}")
            
            self.send_json({'success': True, 'system_prompt': global_system_prompt})
        
        elif path == '/api/settings':
            if 'temperature' in data:
                global_args.temperature = float(data['temperature'])
            if 'max_length' in data:
                global_args.max_length = int(data['max_length'])
            if 'top_k' in data:
                global_args.top_k = int(data['top_k'])
            if 'max_context' in data:
                # Cap at 256k
                max_ctx = int(data['max_context'])
                global_args.max_context = min(max(max_ctx, 1024), 262144)
            self.send_json({'success': True})
        
        elif path == '/api/model/switch':
            from utils.api_chat import save_active_model, get_active_model
            req_mode = data.get('mode', 'local')
            req_model = data.get('model_name', '')
            req_provider = data.get('provider', '')
            current_active = get_active_model()
            if req_mode == 'local':
                save_active_model(
                    mode='local',
                    local_model=req_model or global_model_name or '',
                    api_provider=current_active.get('api_provider', ''),
                    api_model=current_active.get('api_model', '')
                )
            else:
                save_active_model(
                    mode='api',
                    local_model=current_active.get('local_model', global_model_name or ''),
                    api_provider=req_provider,
                    api_model=req_model
                )
            updated = get_active_model()
            self.send_json({'success': True, 'active': updated})
        
        elif path == '/api/api-keys':
            from utils.api_chat import save_api_key, get_api_keys_status
            provider = data.get('provider', '')
            api_key = data.get('api_key', '')
            if not provider:
                self.send_error_json('Missing provider', 400)
            else:
                save_api_key(provider, api_key)
                self.send_json({'success': True, 'keys': get_api_keys_status()})
        
        elif path == '/api/stop':
            global_stop_generation = True
            self.send_json({'success': True})
        
        elif path == '/step_question':
            self.handle_step_question(data)
        
        elif path == '/retry_step':
            self.handle_retry_step(data)
        
        elif path == '/api/tool_call':
            self.handle_direct_tool_call(data)
        
        elif path == '/api/replan':
            self.handle_replan(data)
        
        elif path == '/api/update_plan_analysis':
            conv_id = data.get('conversation_id')
            analysis = data.get('analysis', '')
            if conv_id and conversation_manager.update_plan_analysis(conv_id, analysis):
                self.send_json({'success': True})
            else:
                self.send_error_json('Failed to update plan analysis', 400)
        
        elif path == '/api/execute_code':
            self.handle_execute_code(data)
        
        elif path == '/api/data/upload':
            file_name = data.get('name', 'unknown')
            file_data = data.get('data', '')
            if not file_data:
                self.send_error_json('No file data provided', 400)
                return
            # Sanitize filename
            safe_name = os.path.basename(file_name).replace('..', '_')
            safe_name = re.sub(r'[<>:"/\\|?*]', '_', safe_name)
            if not safe_name:
                safe_name = 'upload'
            # Handle duplicates
            base, ext = os.path.splitext(safe_name)
            dest = os.path.join(UPLOADS_DIR, safe_name)
            counter = 1
            while os.path.exists(dest):
                safe_name = f'{base}_{counter}{ext}'
                dest = os.path.join(UPLOADS_DIR, safe_name)
                counter += 1
            try:
                raw = base64.b64decode(file_data.split(',', 1)[1]) if ',' in file_data else base64.b64decode(file_data)
                with open(dest, 'wb') as f:
                    f.write(raw)
            except Exception as e:
                self.send_error_json(f'Failed to save file: {e}', 500)
                return
            # Extract text for non-image files
            text_content = None
            extraction_method = None
            is_image = ext.lower() in ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp', '.svg', '.ico')
            if not is_image:
                text_content, extraction_method = _extract_text_from_file(dest, safe_name)
                txt_path = dest + '.extracted.txt'
                try:
                    with open(txt_path, 'w', encoding='utf-8') as f:
                        f.write(text_content or '')
                except Exception:
                    pass
            self.send_json({
                'fileName': safe_name,
                'fileSize': os.path.getsize(dest),
                'uploadId': safe_name,
                'textContent': text_content,
                'extractionMethod': extraction_method
            })
        
        else:
            self.send_error_json('Not found', 404)
    
    def do_DELETE(self):
        global global_stop_generation
        parsed = urlparse(self.path)
        path = parsed.path
        
        if path.startswith('/api/conversation/'):
            conv_id = path.split('/')[-1]
            global_stop_generation = True  # Stop generation on delete
            if conversation_manager.delete_conversation(conv_id):
                self.send_json({'success': True})
            else:
                self.send_error_json('Conversation not found', 404)
        elif path.startswith('/api/data/'):
            fname = os.path.basename(path)
            fpath = os.path.join(UPLOADS_DIR, fname)
            if os.path.commonpath([UPLOADS_DIR, os.path.abspath(fpath)]) != UPLOADS_DIR:
                self.send_error_json('Invalid path', 400)
                return
            deleted = False
            if os.path.isfile(fpath):
                os.remove(fpath)
                deleted = True
            txt_path = fpath + '.extracted.txt'
            if os.path.isfile(txt_path):
                os.remove(txt_path)
            if deleted:
                self.send_json({'success': True})
            else:
                self.send_error_json('File not found', 404)
        else:
            self.send_error_json('Not found', 404)
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def handle_execute_code(self, data):
        """Execute Python/R code and capture figures + tables."""
        code = data.get('code', '')
        language = data.get('language', 'python').lower()
        conv_id = data.get('conv_id', 'default')
        step_index = data.get('step_index', 0)

        if not code.strip():
            self.send_json({'success': False, 'error': 'Empty code', 'stdout': '', 'stderr': '', 'figures': [], 'tables': []})
            return

        with model_lock:
            result = _execute_code_subprocess(code, language, conv_id, step_index)

        self.send_json(result)

    def handle_step_question(self, data):
        """Handle question about specific step(s) - stream LLM response."""
        global global_model, global_tokenizer, global_args
        
        conv_id = data.get('conv_id')
        plan_goal = data.get('plan_goal', '')
        plan_steps = data.get('plan_steps', [])
        question = data.get('question', '').strip()
        
        # Support multi-step tags (new format) with backward compatibility
        steps_data = data.get('steps', [])
        if not steps_data and data.get('step_num') is not None:
            steps_data = [{
                'step_num': data.get('step_num'),
                'tool': data.get('tool', ''),
                'step_name': data.get('step_name', ''),
                'step_context': data.get('step_context', ''),
                'previous_steps': data.get('previous_steps', [])
            }]
        
        tagged_nums = [s.get('step_num') for s in steps_data]
        print(f"[step_question] Received question: {question}")
        print(f"[step_question] Tagged steps: {tagged_nums}")
        
        if not question:
            self.send_error_json('Question is required', 400)
            return
        
        if not global_model or not global_tokenizer:
            self.send_error_json('Model not loaded', 500)
            return
        
        # Load conversation and save user question
        conv = conversation_manager.get_conversation(conv_id)
        if not conv:
            conv = conversation_manager.create_conversation(conv_id.split('_')[0] if conv_id else 'chat')
        
        tag_label = ', '.join(f'Step {n}' if n != 0 else 'Plan' for n in tagged_nums)
        user_content = f"[{tag_label} Question] {question}"
        conv['messages'].append({'role': 'user', 'content': user_content})
        user_index = len(conv['messages']) - 1
        conversation_manager.save_conversation(conv_id, conv)
        
        # Build plan structure context (mark tagged steps with arrow)
        tagged_set = set(tagged_nums)
        plan_structure = ""
        if plan_goal:
            plan_structure = f"=== Research Plan ===\nGoal: {plan_goal}\n\nAll Steps:\n"
            for ps in plan_steps:
                marker = "→" if ps.get('num') in tagged_set else " "
                plan_structure += f"{marker} Step {ps.get('num')}: {ps.get('name')} ({ps.get('tool')})\n"
            plan_structure += "\n"
        
        # Build tagged steps context
        tagged_context = "=== Tagged Steps (Question Targets) ===\n"
        for s in steps_data:
            snum = s.get('step_num')
            sname = s.get('step_name', '')
            stool = s.get('tool', '')
            sctx = s.get('step_context', '')[:2000]
            if snum == 0:
                tagged_context += "[Entire Plan]\n\n"
            else:
                tagged_context += f"Step {snum}: {sname} (tool: {stool})\nResult:\n{sctx}\n\n"
            # Include previous steps for this tagged step
            prev_steps = s.get('previous_steps', [])
            if prev_steps:
                tagged_context += "Previous steps:\n"
                for prev in prev_steps:
                    prev_result = prev.get('result', '')[:500]
                    tagged_context += f"  Step {prev.get('num')}: {prev.get('name')} -> {prev_result[:200]}\n"
                tagged_context += "\n"
        
        system_prompt = f"""You are a helpful research assistant. Answer questions about the tagged research steps.

{plan_structure}{tagged_context}
Refer to the research plan and tagged step contexts above to answer the user's question concisely and helpfully.
You may suggest plan modifications or new steps if needed."""
        
        # Set up SSE streaming
        self.send_response(200)
        self.send_header('Content-Type', 'text/event-stream')
        self.send_header('Cache-Control', 'no-cache')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        try:
            # Build history for chat template
            history = [
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': question}
            ]
            
            # Build inputs using existing function
            inputs = build_inputs(history, global_tokenizer, global_args)
            
            # Generate with streaming using existing function
            answer = ""
            for chunk in generate_with_refusal_streaming(global_model, global_tokenizer, inputs, global_args):
                if chunk.get('done', False):
                    break
                    
                token = chunk.get('token', '')
                answer += token
                
                # Send token
                self.wfile.write(f"data: {json.dumps({'token': token})}\n\n".encode('utf-8'))
                self.wfile.flush()
                
                # Check for stop
                if len(answer) > 1000:
                    break
            
            # Save assistant response to history
            conv['messages'].append({'role': 'assistant', 'content': answer})
            assistant_index = len(conv['messages']) - 1
            conversation_manager.save_conversation(conv_id, conv)
            
            # Send done with message indices
            self.wfile.write(f"data: {json.dumps({'done': True, 'user_index': user_index, 'assistant_index': assistant_index})}\n\n".encode('utf-8'))
            self.wfile.flush()
            
        except Exception as e:
            self.wfile.write(f"data: {json.dumps({'error': str(e)})}\n\n".encode('utf-8'))
            self.wfile.flush()
    
    def handle_replan(self, data):
        """Handle re-planning from modified graph structure."""
        conv_id = data.get('conversation_id')
        new_steps = data.get('steps', [])
        goal = data.get('goal', '')
        rerun = data.get('rerun', False)
        
        if not new_steps:
            self.send_error_json('No steps provided', 400)
            return
        
        if not hasattr(self, '_plan_state'):
            self._plan_state = {}
        
        if rerun:
            self._plan_state['all_results'] = []
            self._plan_state['current_step'] = 0
        
        completed_step_ids = set()
        if not rerun and self._plan_state.get('all_results'):
            for r in self._plan_state['all_results']:
                completed_step_ids.add(str(r.get('step_num', '')))
        
        rebuilt_steps = []
        for step in new_steps:
            step_entry = {
                'name': step.get('name', ''),
                'tool': step.get('tool', ''),
                'description': step.get('description', ''),
                'depends_on': step.get('depends_on', []),
                'references': step.get('references', []),
                'inputs': step.get('inputs', {})
            }
            rebuilt_steps.append(step_entry)
        
        self._plan_state['steps'] = rebuilt_steps
        if goal:
            self._plan_state['goal'] = goal
        if not rerun:
            self._plan_state['current_step'] = len(completed_step_ids)
        
        self.send_json({'success': True, 'steps': len(rebuilt_steps), 'completed': len(completed_step_ids)})
    
    def handle_retry_step(self, data):
        """Handle retry of a specific step - LLM regenerates from tool select."""
        global global_model, global_tokenizer, global_args
        
        conv_id = data.get('conv_id')
        step_num = data.get('step_num')
        step_name = data.get('step_name', '')
        original_result = data.get('original_result', '')
        user_edit = data.get('user_edit')
        plan_goal = data.get('plan_goal', '')
        previous_steps = data.get('previous_steps', [])
        
        if not global_model or not global_tokenizer:
            self.send_error_json('Model not loaded', 500)
            return
        
        # Build previous steps context
        previous_steps_context = ""
        if previous_steps:
            previous_steps_context = "=== Previous Step Results ===\n"
            for prev in previous_steps:
                prev_result = prev.get('result', '')[:800]
                previous_steps_context += f"Step {prev.get('num')}: {prev.get('name')}\nResult: {prev_result}\n\n"
        
        # Build system prompt for retry
        system_prompt = f"""You are a research assistant. Regenerate the following research step.

=== Research Plan ===
Goal: {plan_goal}

{previous_steps_context}=== Step to Re-execute ===
Step {step_num}: {step_name}

Previous Result (for reference):
{original_result[:1000]}

{f'User Feedback: {user_edit}' if user_edit else ''}

Re-execute this step referring to the context above.
Select and execute the appropriate tool."""
        
        # Set up SSE streaming
        self.send_response(200)
        self.send_header('Content-Type', 'text/event-stream')
        self.send_header('Cache-Control', 'no-cache')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        try:
            # Build messages for LLM
            messages = [
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': f'Please re-execute Step {step_num}: {step_name}.'}
            ]
            
            # Get tools schema
            tools_schema = get_tools_schema() if TOOLS_AVAILABLE else None
            inputs = build_inputs(messages, global_tokenizer, global_args, tools=tools_schema)
            
            # Stream response
            full_response = ""
            for chunk in generate_with_refusal_streaming(global_model, global_tokenizer, inputs, global_args):
                if chunk.get('done', False):
                    break
                
                token = chunk.get('token', '')
                full_response += token
                self.wfile.write(f"data: {json.dumps({'token': token})}\n\n".encode('utf-8'))
                self.wfile.flush()
            
            # Check for tool calls and execute
            if TOOLS_AVAILABLE and detect_tool_call(full_response):
                remaining_text, tool_calls = parse_tool_calls(full_response)
                
                for call in tool_calls:
                    # Send tool call event
                    self.wfile.write(f"data: {json.dumps({'tool_call': {'name': call['name'], 'arguments': call['arguments'], 'status': 'running'}})}\n\n".encode('utf-8'))
                    self.wfile.flush()
                    
                    # Execute tool
                    result = execute_tool_call(call['name'], call['arguments'])
                    
                    # Send tool result (include tool name for frontend step matching)
                    self.wfile.write(f"data: {json.dumps({'tool_result': {**result, 'tool': call['name']}})}\n\n".encode('utf-8'))
                    self.wfile.flush()
            
            # Send done
            self.wfile.write(f"data: {json.dumps({'done': True})}\n\n".encode('utf-8'))
            self.wfile.flush()
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.wfile.write(f"data: {json.dumps({'error': str(e)})}\n\n".encode('utf-8'))
            self.wfile.flush()
    
    def handle_direct_tool_call(self, data):
        """Handle direct tool call from Detail Panel.
        
        This endpoint allows the frontend to directly call tools (like analyze_plan, code_gen)
        without going through the full chat flow.
        """
        tool_name = data.get('tool')
        args = data.get('args', {})
        
        if not tool_name:
            self.send_error_json('Tool name is required', 400)
            return
        
        if not TOOLS_AVAILABLE:
            self.send_error_json('Tools not available', 500)
            return
        
        try:
            from tools.base import get_tool
            tool = get_tool(tool_name)
            
            if not tool:
                self.send_error_json(f'Tool "{tool_name}" not found', 404)
                return
            
            with model_lock:
                result = tool.execute(**args)
            
            self.send_json({
                'success': result.get('success', True),
                'tool': tool_name,
                'result': result.get('result', result)
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.send_error_json(f'Tool execution failed: {str(e)}', 500)
    
    def handle_chat(self, data):
        """Handle chat message with SSE streaming."""
        conv_id = data.get('conversation_id')
        message = data.get('message', '').strip()
        files = data.get('files', [])
        is_rerun = data.get('rerun', False)
        rerun_steps = data.get('rerun_steps', [])
        rerun_goal = data.get('rerun_goal', '')
        
        if not conv_id or (not message and not files and not is_rerun):
            self.send_error_json('Missing conversation_id or message')
            return
        
        # Get conversation
        conv = conversation_manager.get_conversation(conv_id)
        if not conv:
            self.send_error_json('Conversation not found', 404)
            return
        
        # Process attached files
        processed_files = []
        file_descriptions = []
        
        for f in files:
            file_type = f.get('type')
            file_name = f.get('name', 'unknown')

            if file_type == 'image':
                upload_id = f.get('uploadId')
                if upload_id:
                    processed_files.append({
                        'type': 'image',
                        'name': file_name,
                        'uploadId': upload_id
                    })
                else:
                    processed_files.append({
                        'type': 'image',
                        'name': file_name,
                        'data': f.get('data', '')
                    })
                file_descriptions.append(f"[Image: {file_name}]")
            elif file_type == 'document':
                text_content = f.get('textContent', '')
                processed_files.append({
                    'type': 'document',
                    'name': file_name,
                    'uploadId': f.get('uploadId'),
                    'textContent': text_content
                })
                file_descriptions.append(f"[Document: {file_name}]")
            elif file_type == 'audio':
                processed_files.append({
                    'type': 'audio',
                    'name': file_name,
                    'data': f.get('data', '')
                })
                file_descriptions.append(f"[Audio: {file_name}]")
        
        # Build full message content with document context
        doc_context = []
        for f in processed_files:
            if f['type'] == 'document' and f.get('textContent'):
                doc_context.append(f"--- {f['name']} ---\n{f['textContent']}")

        full_message = message
        if doc_context:
            full_message = '\n\n'.join(doc_context) + '\n\n' + full_message
        if file_descriptions:
            file_header = ' '.join(file_descriptions)
            full_message = file_header + ('\n' + full_message if full_message else '')
        
        # Skip saving user message for rerun (no actual user message)
        if not is_rerun:
            conversation_manager.add_message(conv_id, 'user', full_message, files=processed_files)
        
        # Send SSE headers
        self.send_response(200)
        self.send_header('Content-Type', 'text/event-stream')
        self.send_header('Cache-Control', 'no-cache')
        self.send_header('Connection', 'keep-alive')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        # Initialize outside try block so except can access them
        full_response = ""
        response_saved = False
        
        # Check if we should use API mode
        from utils.api_chat import get_active_model as _get_active
        _active_cfg = _get_active()
        _use_api_mode = (
            _active_cfg.get('mode') == 'api'
            and _active_cfg.get('api_provider')
            and _active_cfg.get('api_model')
        )
        
        if _use_api_mode and not is_rerun:
            # --- API Model Path (not used for rerun - rerun requires tool execution) ---
            try:
                from utils.api_chat import stream_chat
                conv = conversation_manager.get_conversation(conv_id)
                history = []
                for msg in conv.get('messages', [])[:-1]:
                    history.append({'role': msg['role'], 'content': msg['content']})
                
                provider = _active_cfg.get('api_provider', '')
                api_model = _active_cfg.get('api_model', '')
                system_prompt = global_system_prompt or ''
                
                print(f"[DEBUG] API mode: provider={provider}, model={api_model}")
                print(f"[DEBUG] system_prompt ({len(system_prompt)} chars): {system_prompt[:80]}{'...' if len(system_prompt) > 80 else ''}")
                
                for chunk in stream_chat(
                    provider=provider,
                    model=api_model,
                    history=history,
                    message=full_message,
                    system_prompt=system_prompt,
                    temperature=global_args.temperature,
                    max_tokens=global_args.max_length
                ):
                    if chunk.get('error'):
                        self.wfile.write(f"data: {json.dumps({'error': chunk['error']})}\n\n".encode('utf-8'))
                        self.wfile.flush()
                        break
                    if chunk.get('token'):
                        full_response += chunk['token']
                        self.wfile.write(f"data: {json.dumps({'token': chunk['token']})}\n\n".encode('utf-8'))
                        self.wfile.flush()
                    if chunk.get('done'):
                        break
                
                if full_response:
                    conversation_manager.add_message(conv_id, 'assistant', full_response)
                    response_saved = True
                
                self.wfile.write(f"data: {json.dumps({'done': True})}\n\n".encode('utf-8'))
                self.wfile.flush()
            except Exception as e:
                import traceback
                traceback.print_exc()
                error_msg = f"API chat error: {str(e)}"
                if not response_saved and full_response:
                    conversation_manager.add_message(conv_id, 'assistant', full_response)
                try:
                    self.wfile.write(f"data: {json.dumps({'error': error_msg})}\n\n".encode('utf-8'))
                    self.wfile.flush()
                except:
                    pass
            return
        
        # --- Local Model Path (original logic) ---
        try:
            global global_stop_generation
            global_stop_generation = False  # Reset stop flag
            
            # Build messages for model
            conv = conversation_manager.get_conversation(conv_id)
            history = []
            msgs = conv.get('messages', [])
            if not is_rerun:
                msgs = msgs[:-1]  # Exclude the just-added user message
            for msg in msgs:
                history.append({'role': msg['role'], 'content': msg['content']})
            
            if is_rerun and rerun_steps:
                # --- RERUN BYPASS: skip LLM create_plan, use graph steps directly ---
                history = [m for m in history if '[PLAN_COMPLETE]' not in m.get('content', '')]
                
                self._plan_state = {
                    'steps': rerun_steps,
                    'goal': rerun_goal,
                    'current_step': 0,
                    'all_results': []
                }
                
                current_system_prompt = load_tool_select_prompt()
                if TOOLS_AVAILABLE:
                    tools_schema = get_tools_schema()
                else:
                    tools_schema = None
                
                step = rerun_steps[0]
                ref_ctx = _build_ref_context(step)
                step_context = f"Execute step 1: {step.get('name', '')}. {step.get('description', '')}{ref_ctx} Choose and call the appropriate tool(s)."
                history.append({'role': 'user', 'content': step_context})
                
                step_start_evt = json.dumps({'step_start': {'step': 1}})
                self.wfile.write(f"data: {step_start_evt}\n\n".encode('utf-8'))
                self.wfile.flush()
                
                print(f"[DEBUG] Rerun: bypassing create_plan, {len(rerun_steps)} steps from graph")
                
                synthetic_plan_msg = f"[TOOL_CALLS]create_plan[ARGS]{json.dumps({'goal': rerun_goal, 'steps': rerun_steps}, ensure_ascii=False)}"
                conversation_manager.add_message(conv_id, 'assistant', synthetic_plan_msg)
                response_saved = True
                
                inputs = _build_step_inputs(step, history, current_system_prompt, global_tokenizer, global_args, tools_schema)
            else:
                # --- Normal flow ---
                # Extract images from processed files
                images = []
                for f in processed_files:
                    if f['type'] == 'image':
                        try:
                            if f.get('uploadId'):
                                img_path = os.path.join(UPLOADS_DIR, f['uploadId'])
                                if os.path.isfile(img_path):
                                    img = Image.open(img_path).convert('RGB')
                                    images.append(img)
                                else:
                                    print(f"[Warning] Upload file not found: {f['uploadId']}")
                            elif f.get('data'):
                                img = process_image_from_base64(f['data'])
                                images.append(img)
                        except Exception as e:
                            print(f"[Warning] Failed to process image: {e}")
                
                # Mode-based tool routing
                mode = data.get('mode', 'agent')
                use_tools = True
                current_system_prompt = global_system_prompt
                
                if mode == 'agent':
                    # Agent mode: no tool routing, plain chat
                    use_tools = False
                    tools_schema = None
                    print(f"[DEBUG] Agent mode: tools disabled")
                else:
                    # Plan mode: load plan system prompt and enable tools
                    plan_prompt_path = os.path.join(os.path.dirname(__file__), "prompts", "PLAN_SYSTEM_PROMPT.txt")
                    if os.path.exists(plan_prompt_path):
                        with open(plan_prompt_path, "r", encoding="utf-8") as f:
                            current_system_prompt = f.read().strip()
                    if TOOLS_AVAILABLE:
                        from tools.base import get_plan_schema
                        tools_schema = get_plan_schema()
                    else:
                        tools_schema = None
                
                # Build inputs (multimodal if images present)
                sp_preview = (current_system_prompt or '')[:80]
                print(f"[DEBUG] Local {mode} mode, system_prompt ({len(current_system_prompt or '')} chars): {sp_preview}{'...' if len(current_system_prompt or '') > 80 else ''}")
                if images and global_processor is not None:
                    messages = build_multimodal_messages(history, message, images, current_system_prompt)
                    inputs = build_inputs_multimodal(messages, images, global_processor, global_tokenizer, global_args)
                else:
                    messages = build_messages(history, message, current_system_prompt)
                    inputs = build_inputs(messages, global_tokenizer, global_args, tools=tools_schema, inject_tool_call_prefix=False)
            
            # Stream response with tool call support
            stopped = False
            max_tool_iterations = 10  # Prevent infinite loops
            tool_iteration = 0
            
            model_lock.acquire()
            while tool_iteration < max_tool_iterations:
                tool_iteration += 1
                
                for chunk in generate_with_refusal_streaming(
                    global_model, global_tokenizer, inputs, global_args
                ):
                    # Check stop flag
                    if global_stop_generation:
                        stopped = True
                        break
                    
                    if chunk.get('token'):
                        token = chunk['token']
                        full_response += token
                        self.wfile.write(f"data: {json.dumps({'token': token})}\n\n".encode('utf-8'))
                        self.wfile.flush()
                    
                    if chunk.get('done'):
                        break
                
                if stopped:
                    break
                
                # Debug: Log full response content
                print(f"[DEBUG] Full response length: {len(full_response)}")
                print(f"[DEBUG] Full response preview: {full_response[:500]}")
                print(f"[DEBUG] Contains [TOOL_CALLS]: {'[TOOL_CALLS]' in full_response}")
                print(f"[DEBUG] Contains [ARGS]: {'[ARGS]' in full_response}")
                
                # Check for tool calls in the response (using adapter-aware detection)
                detected = detect_tool_call(full_response) if TOOLS_AVAILABLE else False
                print(f"[DEBUG] detect_tool_call result: {detected}")
                
                # Check for malformed tool call (starts with [ARGS] instead of [TOOL_CALLS])
                malformed_tool_call = not detected and '[ARGS]' in full_response and TOOLS_AVAILABLE
                
                if TOOLS_AVAILABLE and (detected or malformed_tool_call):
                    # Handle malformed tool call - retry with correct format instruction
                    if malformed_tool_call:
                        print(f"[DEBUG] Malformed tool call detected (starts with [ARGS]), requesting retry...")
                        
                        for retry_num, response in retry_llm_generation(
                            history, full_response,
                            "Your tool call format is wrong. You MUST call create_plan with [TOOL_CALLS] prefix.\n"
                            "Correct format: [TOOL_CALLS]create_plan[ARGS]{\"goal\": \"...\", \"steps\": [...]}\n"
                            "Output ONLY the create_plan tool call with correct format. Do NOT call other tools directly.",
                            current_system_prompt, global_tokenizer, global_model, global_args, tools_schema
                        ):
                            print(f"[DEBUG] Malformed tool call retry {retry_num}/{MAX_RETRY_ATTEMPTS}")
                            detected = detect_tool_call(response)
                            print(f"[DEBUG] After malformed retry {retry_num}, detect_tool_call result: {detected}")
                            if detected:
                                full_response = response
                                break
                        
                        if not detected:
                            print(f"[DEBUG] Malformed tool call retry failed after {MAX_RETRY_ATTEMPTS} attempts")
                    
                    if detected:
                        remaining_text, tool_calls = parse_tool_calls(full_response)
                        
                        # Parsing failed (JSON incomplete) - retry up to 3 times
                        parse_retry_count = 0
                        while not tool_calls and parse_retry_count < 3:
                            parse_retry_count += 1
                            print(f"[DEBUG] Tool call parsing failed (JSON incomplete), retry {parse_retry_count}/3")
                            
                            # Ask LLM to regenerate properly formatted tool call
                            retry_history = history.copy()
                            retry_history.append({'role': 'assistant', 'content': full_response})
                            retry_history.append({'role': 'user', 'content': 
                                "Your tool call was malformed or incomplete. Please output the tool call again with valid JSON. "
                                "Format: [TOOL_CALLS]tool_name[ARGS]{\"arg\": \"value\"}"
                            })
                            
                            # Generate retry
                            retry_messages = build_messages(retry_history, "", current_system_prompt)
                            retry_inputs = build_inputs(retry_messages, global_tokenizer, global_args, tools=tools_schema)
                            
                            retry_response = ""
                            for chunk in generate_with_refusal_streaming(global_model, global_tokenizer, retry_inputs, global_args):
                                if chunk.get('token'):
                                    retry_response += chunk['token']
                                if chunk.get('done'):
                                    break
                            
                            # Try parsing again
                            if detect_tool_call(retry_response):
                                full_response = retry_response
                                remaining_text, tool_calls = parse_tool_calls(retry_response)
                                if tool_calls:
                                    print(f"[DEBUG] Tool call parsing succeeded on retry {parse_retry_count}")
                            else:
                                print(f"[DEBUG] Retry {parse_retry_count} did not produce a tool call")
                        
                        if tool_calls:
                            # Get adapter for proper formatting
                            adapter = get_adapter_for_model(global_args.model_type) if TOOLS_AVAILABLE else None
                            
                            # Execute tool calls and send results to UI
                            tool_results = []
                            
                            for _tc_idx, call in enumerate(tool_calls):
                                _tc_remaining = len(tool_calls) - _tc_idx - 1
                                tool_name = call.get('name', '')
                                
                                # Check if tool is None or empty - retry to get valid tool, or fallback to LLM text
                                if not tool_name or tool_name.lower() == 'none':
                                    print(f"[DEBUG] Tool is None/empty, attempting to get valid tool...")
                                    valid_tool_found = False
                                    
                                    # Get available tools dynamically from tools_schema
                                    available_tools = ", ".join([t.get('function', {}).get('name', t.get('name', '')) for t in tools_schema]) if tools_schema else "no tools available"
                                    
                                    for retry_num, retry_response in retry_llm_generation(
                                        history, full_response,
                                        f"You selected tool 'None' which is not valid. "
                                        f"Please select an appropriate tool from: {available_tools}. "
                                        "Output ONLY the tool call with correct format: [TOOL_CALLS]tool_name[ARGS]{...}",
                                        current_system_prompt, global_tokenizer, global_model, global_args, tools_schema
                                    ):
                                        print(f"[DEBUG] No-tool retry {retry_num}/{MAX_RETRY_ATTEMPTS}")
                                        _, retry_tool_calls = parse_tool_calls(retry_response)
                                        if retry_tool_calls:
                                            retry_call = retry_tool_calls[0]
                                            retry_tool_name = retry_call.get('name', '')
                                            if retry_tool_name and retry_tool_name.lower() != 'none':
                                                print(f"[DEBUG] Valid tool selected on retry: {retry_tool_name}")
                                                call = retry_call
                                                tool_name = retry_tool_name
                                                valid_tool_found = True
                                                break
                                        print(f"[DEBUG] Retry {retry_num}: Still no valid tool")
                                    
                                    # Fallback to LLM text generation if no valid tool after retries
                                    if not valid_tool_found:
                                        print(f"[DEBUG] No valid tool after {MAX_RETRY_ATTEMPTS} retries, falling back to LLM text generation")
                                        
                                        # Generate text response instead of tool call
                                        text_history = history.copy()
                                        step_info = ""
                                        if hasattr(self, '_plan_state') and self._plan_state.get('steps'):
                                            step_idx = self._plan_state.get('current_step', 0)
                                            if step_idx < len(self._plan_state['steps']):
                                                step = self._plan_state['steps'][step_idx]
                                                step_info = f"For step: {step.get('name', '')}\n{step.get('description', '')}\n\n"
                                        
                                        text_history.append({'role': 'user', 'content': 
                                            f"{step_info}There is no appropriate tool for this task. "
                                            "Please provide a detailed text response explaining how to accomplish this step."
                                        })
                                        
                                        text_messages = build_messages(text_history, "", global_system_prompt)
                                        text_inputs = build_inputs(text_messages, global_tokenizer, global_args, tools=None)
                                        
                                        llm_text_response = ""
                                        for chunk in generate_with_refusal_streaming(global_model, global_tokenizer, text_inputs, global_args):
                                            if chunk.get('token'):
                                                llm_text_response += chunk['token']
                                                # Stream to UI
                                                self.wfile.write(f"data: {json.dumps({'token': chunk['token']})}\n\n".encode('utf-8'))
                                                self.wfile.flush()
                                            if chunk.get('done'):
                                                break
                                        
                                        # Create a fake successful result with LLM text
                                        result = {
                                            'success': True,
                                            'result': {'text': llm_text_response, 'source': 'llm_fallback'},
                                            'tool': 'llm_text'
                                        }
                                        call = {'id': call.get('id', 'llm_fallback'), 'name': 'llm_text', 'arguments': {}}
                                        tool_results.append({'call': call, 'result': result})
                                        
                                        # Send tool result event
                                        current_step = self._plan_state.get('current_step', 0) + 1 if hasattr(self, '_plan_state') and self._plan_state.get('steps') else None
                                        step_info_dict = {'step': current_step} if current_step else {}
                                        self.wfile.write(f"data: {json.dumps({'tool_result': {**result, 'tool': 'llm_text', **step_info_dict, 'tools_remaining': _tc_remaining}})}\n\n".encode('utf-8'))
                                        self.wfile.flush()
                                        continue  # Skip normal tool execution
                                
                                # Send tool call event to UI
                                self.wfile.write(f"data: {json.dumps({'tool_call': {'name': call['name'], 'arguments': call['arguments'], 'status': 'running'}})}\n\n".encode('utf-8'))
                                self.wfile.flush()
                                
                                # Pass conv_id/step_index and file refs to code_gen
                                if call['name'] == 'code_gen' and hasattr(self, '_plan_state') and self._plan_state.get('steps'):
                                    call['arguments']['conv_id'] = conv_id
                                    step_idx = self._plan_state.get('current_step', 0)
                                    call['arguments']['step_index'] = step_idx
                                    step = self._plan_state['steps'][step_idx]
                                    file_refs = _collect_file_refs(step)
                                    if file_refs:
                                        existing_ctx = call['arguments'].get('context', '')
                                        call['arguments']['context'] = (existing_ctx +
                                            '\n\nAvailable data files: ' + ', '.join(file_refs)).strip()
                                    
                                    # Save previous step results so subprocess code can access them
                                    prev_results = self._plan_state.get('all_results', [])
                                    if prev_results:
                                        out_dir = os.path.join(OUTPUTS_DIR, str(conv_id), f'step_{step_idx}')
                                        out_dir_fwd = out_dir.replace('\\', '/')
                                        os.makedirs(out_dir, exist_ok=True)
                                        prev_data_path = os.path.join(out_dir, '_prev_data.json')
                                        try:
                                            with open(prev_data_path, 'w', encoding='utf-8') as pf:
                                                json.dump(prev_results, pf, ensure_ascii=False, default=str)
                                            print(f"[code_gen] Saved {len(prev_results)} previous results to {prev_data_path}")
                                        except Exception as e:
                                            print(f"[code_gen] Failed to save prev data: {e}")
                                        call['arguments']['data_dir'] = out_dir_fwd
                                
                                # Execute the tool
                                result = execute_tool_call(call['name'], call['arguments'])
                                
                                # If failed, ask LLM to regenerate the tool call using unified retry function
                                # Skip retry for code_gen -- it has its own internal fix loop
                                if not result.get('success', False) and call['name'] != 'code_gen':
                                    error_msg = result.get('error', 'Unknown error')
                                    print(f"[DEBUG] Tool {call['name']} failed: {error_msg}")
                                    
                                    for retry_num, retry_response in retry_llm_generation(
                                        history, full_response,
                                        f"Tool call failed with error: {error_msg}\n\nPlease try again with correct arguments. Output ONLY the tool call, no explanation.",
                                        current_system_prompt, global_tokenizer, global_model, global_args, tools_schema
                                    ):
                                        print(f"[DEBUG] Tool retry {retry_num}/{MAX_RETRY_ATTEMPTS}")
                                        
                                        # Parse new tool call from retry response
                                        _, retry_tool_calls = parse_tool_calls(retry_response)
                                        if retry_tool_calls:
                                            retry_call = retry_tool_calls[0]
                                            print(f"[DEBUG] Regenerated tool call: {retry_call['name']}")
                                            print(f"[DEBUG] New arguments: {retry_call['arguments']}")
                                            
                                            # Update UI with new tool call
                                            self.wfile.write(f"data: {json.dumps({'tool_call': {'name': retry_call['name'], 'arguments': retry_call['arguments'], 'status': 'running'}})}\n\n".encode('utf-8'))
                                            self.wfile.flush()
                                            
                                            # Execute regenerated tool call
                                            result = execute_tool_call(retry_call['name'], retry_call['arguments'])
                                            call = retry_call  # Update call reference for result tracking
                                            
                                            if result.get('success', False):
                                                print(f"[DEBUG] Tool call succeeded on retry {retry_num}!")
                                                break
                                        else:
                                            print(f"[DEBUG] Retry {retry_num}: Failed to parse regenerated tool call")
                                    
                                    if not result.get('success', False):
                                        print(f"[DEBUG] Tool call failed after {MAX_RETRY_ATTEMPTS} retries")
                                
                                tool_results.append({'call': call, 'result': result})
                                
                                # Send tool result event to UI (include tool name and step number for frontend step matching)
                                current_step = self._plan_state.get('current_step', 0) + 1 if hasattr(self, '_plan_state') and self._plan_state.get('steps') else None
                                step_info = {'step': current_step} if current_step else {}
                                self.wfile.write(f"data: {json.dumps({'tool_result': {**result, 'tool': call['name'], **step_info, 'tools_remaining': _tc_remaining}})}\n\n".encode('utf-8'))
                                self.wfile.flush()
                            
                            # Add assistant message with tool call SUMMARY to history (not raw tool call text)
                            # This avoids caching unnecessary tool call tokens in KV cache
                            tool_names = [call['name'] for call in tool_calls]
                            tool_summary = f"Called tool(s): {', '.join(tool_names)}"
                            history.append({'role': 'assistant', 'content': tool_summary})
                            
                            # Check if create_plan was executed - switch to TOOL_SELECT_PROMPT
                            executed_tool_name = tool_calls[0]['name'] if tool_calls else None
                            if executed_tool_name == 'create_plan':
                                # Extract plan steps for subsequent tool selection
                                plan_result = tool_results[0]['result'] if tool_results else {}
                                plan_steps = plan_result.get('result', {}).get('steps', [])
                                plan_goal = plan_result.get('result', {}).get('goal', '')
                                if plan_steps:
                                    # Store plan state for step execution
                                    if not hasattr(self, '_plan_state'):
                                        self._plan_state = {}
                                    self._plan_state['steps'] = plan_steps
                                    self._plan_state['goal'] = plan_goal
                                    self._plan_state['current_step'] = 0
                                    self._plan_state['all_results'] = []  # Initialize results accumulator
                                    # Load TOOL_SELECT_PROMPT for next iteration
                                    current_system_prompt = load_tool_select_prompt()
                                    # Add step context for first step
                                    step = plan_steps[0]
                                    ref_ctx = _build_ref_context(step)
                                    step_context = f"Execute step 1: {step.get('name', '')}. {step.get('description', '')}{ref_ctx} Choose and call the appropriate tool(s)."
                                    history.append({'role': 'user', 'content': step_context})
                                    
                                    # Send step_start event to UI (tool will be determined at execution time)
                                    self.wfile.write(f"data: {json.dumps({'step_start': {'step': 1}})}\n\n".encode('utf-8'))
                                    self.wfile.flush()
                                    # Save the create_plan tool call so the plan can be restored if the user navigates away
                                    if full_response.strip():
                                        conversation_manager.add_message(conv_id, 'assistant', full_response.strip())
                                    response_saved = True
                            else:
                                # Normal tool execution - check if we're in plan execution mode
                                if hasattr(self, '_plan_state') and self._plan_state.get('steps'):
                                    current_system_prompt = load_tool_select_prompt()
                                    
                                    # Accumulate this tool's result (store full data for UI display)
                                    step_idx = self._plan_state['current_step']
                                    for item in tool_results:
                                        call = item['call']
                                        result = item['result']
                                        
                                        # Store full result data including thought, action, and result
                                        res_data = result.get('result', {})
                                        result_entry = {
                                            'step': step_idx + 1,
                                            'tool': call['name'],
                                            'success': result.get('success', False),
                                            'thought': result.get('thought', ''),
                                            'action': result.get('action', ''),
                                            'error': result.get('error', ''),
                                            'result': res_data if isinstance(res_data, dict) else {'title': str(res_data)}
                                        }
                                        if global_stop_generation:
                                            result_entry['stopped'] = True
                                        self._plan_state['all_results'].append(result_entry)
                                    
                                    # Move to next step
                                    self._plan_state['current_step'] += 1
                                    step_idx = self._plan_state['current_step']
                                    steps = self._plan_state['steps']
                                    if step_idx < len(steps):
                                        step = steps[step_idx]
                                        ref_ctx = _build_ref_context(step)
                                        step_context = f"Execute step {step_idx + 1}: {step.get('name', '')}. {step.get('description', '')}{ref_ctx} Choose and call the appropriate tool(s)."
                                        tool_context = "\n\n".join([format_tool_result_for_llm(item['result']) for item in tool_results])
                                        history.append({'role': 'user', 'content': f"Tool result:\n{tool_context}\n\n{step_context}"})
                                        
                                        # Send step_start event to UI (tool will be determined at execution time)
                                        self.wfile.write(f"data: {json.dumps({'step_start': {'step': step_idx + 1}})}\n\n".encode('utf-8'))
                                        self.wfile.flush()
                                    else:
                                        # All steps completed - save plan results and send done
                                        plan_complete_data = {
                                            'goal': self._plan_state.get('goal', ''),
                                            'steps': self._plan_state.get('steps', []),
                                            'results': self._plan_state.get('all_results', [])
                                        }
                                        plan_complete_msg = f"[PLAN_COMPLETE]{json.dumps(plan_complete_data, ensure_ascii=False)}"
                                        if not conversation_manager.replace_last_plan_message(conv_id, plan_complete_msg):
                                            conversation_manager.add_message(conv_id, 'assistant', plan_complete_msg)
                                        response_saved = True
                                        
                                        self._plan_state = {}
                                        
                                        # Send done event WITH plan results
                                        self.wfile.write(f"data: {json.dumps({'done': True, 'plan_complete': plan_complete_data})}\n\n".encode('utf-8'))
                                        self.wfile.flush()
                                        break
                                else:
                                    current_system_prompt = global_system_prompt
                                    # Add tool results using adapter format (or fallback)
                                    if adapter:
                                        for item in tool_results:
                                            call = item['call']
                                            result = item['result']
                                            tool_result_obj = ToolResult(
                                                call_id=call['id'],
                                                name=call['name'],
                                                content=result.get('result', result),
                                                success=result.get('success', True)
                                            )
                                            history.append(adapter.format_tool_result(tool_result_obj))
                                    else:
                                        # Legacy fallback
                                        tool_context = "\n\n".join([format_tool_result_for_llm(item['result']) for item in tool_results])
                                        history.append({'role': 'user', 'content': f"Tool results:\n{tool_context}\n\nNow call the next tool. Output ONLY a tool call, no text."})
                        
                        # Save tool call response to DB before resetting
                        # Skip during plan execution - intermediate tool calls are not user-facing
                        if full_response.strip() and not response_saved:
                            if not (hasattr(self, '_plan_state') and self._plan_state.get('steps')):
                                conversation_manager.add_message(conv_id, 'assistant', full_response.strip())
                                response_saved = True
                        
                        # Rebuild inputs for next iteration (with tools schema)
                        full_response = ""  # Reset for next generation
                        response_saved = False  # Allow saving the next iteration's response
                        tools_schema = get_tools_schema() if TOOLS_AVAILABLE else None
                        # Use multimodal inputs if current step has image data
                        if hasattr(self, '_plan_state') and self._plan_state.get('steps'):
                            current_step_idx = self._plan_state.get('current_step', 0)
                            current_step_data = self._plan_state['steps'][current_step_idx] if current_step_idx < len(self._plan_state['steps']) else {}
                            inputs = _build_step_inputs(current_step_data, history, current_system_prompt, global_tokenizer, global_args, tools_schema)
                        else:
                            messages = build_messages(history, "", current_system_prompt)
                            inputs = build_inputs(messages, global_tokenizer, global_args, tools=tools_schema)
                        
                        continue  # Continue the while loop for next generation
                
                # No tool call detected - check if we're in plan execution
                if hasattr(self, '_plan_state') and self._plan_state.get('steps'):
                    print(f"[DEBUG] No tool call during plan execution, attempting retry...")
                    
                    # Retry to get tool call
                    tool_obtained = False
                    for retry_num, retry_response in retry_llm_generation(
                        history, full_response,
                        "You MUST output a tool call. Format: [TOOL_CALLS]tool_name[ARGS]{...}\n"
                        "Do NOT output explanations. Output ONLY the tool call.",
                        current_system_prompt, global_tokenizer, global_model, global_args, tools_schema
                    ):
                        print(f"[DEBUG] No-tool retry {retry_num}/{MAX_RETRY_ATTEMPTS}")
                        if detect_tool_call(retry_response):
                            _, retry_tool_calls = parse_tool_calls(retry_response)
                            if retry_tool_calls:
                                # Success - proceed to tool execution
                                full_response = retry_response
                                detected = True
                                tool_obtained = True
                                print(f"[DEBUG] Tool call obtained on retry {retry_num}")
                                break
                    
                    if tool_obtained:
                        # Re-enter the tool processing block
                        continue
                    
                    # Fallback: save text as step result and move to next step
                    print(f"[DEBUG] Using text response as step result (fallback)")
                    step_idx = self._plan_state['current_step']
                    fallback_entry = {
                        'step': step_idx + 1,
                        'tool': 'text_fallback',
                        'success': True,
                        'result': {'text': full_response}
                    }
                    if global_stop_generation:
                        fallback_entry['stopped'] = True
                    self._plan_state['all_results'].append(fallback_entry)
                    
                    # Send fallback result to UI
                    self.wfile.write(f"data: {json.dumps({'tool_result': {'success': True, 'tool': 'text_fallback', 'result': {'text': full_response}, 'step': step_idx + 1}})}\n\n".encode('utf-8'))
                    self.wfile.flush()
                    
                    # Move to next step
                    self._plan_state['current_step'] += 1
                    step_idx = self._plan_state['current_step']
                    steps = self._plan_state['steps']
                    
                    if step_idx < len(steps):
                        step = steps[step_idx]
                        suggested_tool = step.get('tool', '')
                        tool_hint = f" (suggested: {suggested_tool})" if suggested_tool else ""
                        ref_ctx = _build_ref_context(step)
                        step_context = f"Execute step {step_idx + 1}: {step.get('name', '')}{tool_hint}. {step.get('description', '')}{ref_ctx} Call the appropriate tool(s) needed."
                        history.append({'role': 'user', 'content': step_context})
                        
                        # Notify UI of step start
                        self.wfile.write(f"data: {json.dumps({'step_start': {'step': step_idx + 1, 'tool': step.get('tool', '')}})}\n\n".encode('utf-8'))
                        self.wfile.flush()
                        
                        full_response = ""
                        inputs = _build_step_inputs(step, history, current_system_prompt, global_tokenizer, global_args, tools_schema)
                        continue  # Next generation
                    else:
                        # All steps completed - send plan_complete
                        plan_complete_data = {
                            'goal': self._plan_state.get('goal', ''),
                            'steps': self._plan_state.get('steps', []),
                            'results': self._plan_state.get('all_results', [])
                        }
                        plan_complete_msg = f"[PLAN_COMPLETE]{json.dumps(plan_complete_data, ensure_ascii=False)}"
                        if not conversation_manager.replace_last_plan_message(conv_id, plan_complete_msg):
                            conversation_manager.add_message(conv_id, 'assistant', plan_complete_msg)
                        response_saved = True
                        self._plan_state = {}
                        
                        self.wfile.write(f"data: {json.dumps({'done': True, 'plan_complete': plan_complete_data})}\n\n".encode('utf-8'))
                        self.wfile.flush()
                        break
                
                # No more tool calls, we're done
                break
            model_lock.release()
            
            # Save final response
            # Skip during plan execution - intermediate tool calls are not user-facing
            if full_response.strip() and not response_saved:
                if not (hasattr(self, '_plan_state') and self._plan_state.get('steps')):
                    conversation_manager.add_message(conv_id, 'assistant', full_response.strip())
                    response_saved = True
                    self.wfile.write(f"data: {json.dumps({'done': True})}\n\n".encode('utf-8'))
                    self.wfile.flush()
            
            # If stopped, save partial plan results or partial response
            if stopped:
                if hasattr(self, '_plan_state') and self._plan_state.get('steps'):
                    partial_data = {
                        'goal': self._plan_state.get('goal', ''),
                        'steps': self._plan_state.get('steps', []),
                        'results': self._plan_state.get('all_results', []),
                        'stopped': True
                    }
                    partial_msg = f"[PLAN_COMPLETE]{json.dumps(partial_data, ensure_ascii=False)}"
                    if not conversation_manager.replace_last_plan_message(conv_id, partial_msg):
                        conversation_manager.add_message(conv_id, 'assistant', partial_msg)
                    response_saved = True
                    self._plan_state = {}
                elif full_response.strip() and not response_saved:
                    conversation_manager.add_message(conv_id, 'assistant', full_response.strip())
                    response_saved = True
                try:
                    self.wfile.write(f"data: {json.dumps({'done': True, 'stopped': True})}\n\n".encode('utf-8'))
                    self.wfile.flush()
                except:
                    pass  # Client disconnected, but response is already saved
        
        except Exception as e:
            if model_lock._is_owned():
                model_lock.release()
            
            # During plan execution, save partial plan results instead of raw tool call
            if hasattr(self, '_plan_state') and self._plan_state.get('steps'):
                partial_data = {
                    'goal': self._plan_state.get('goal', ''),
                    'steps': self._plan_state.get('steps', []),
                    'results': self._plan_state.get('all_results', []),
                    'stopped': True
                }
                partial_msg = f"[PLAN_COMPLETE]{json.dumps(partial_data, ensure_ascii=False)}"
                if not response_saved:
                    if not conversation_manager.replace_last_plan_message(conv_id, partial_msg):
                        conversation_manager.add_message(conv_id, 'assistant', partial_msg)
                    response_saved = True
                self._plan_state = {}
            elif full_response and full_response.strip() and not response_saved:
                conversation_manager.add_message(conv_id, 'assistant', full_response.strip())
                print(f"[INFO] Saved partial response ({len(full_response)} chars) after error")
            
            error_msg = str(e)
            # Only log non-connection errors
            if 'Broken pipe' not in error_msg and 'Connection' not in error_msg:
                print(f"\n[ERROR] Exception in handle_chat: {type(e).__name__}: {error_msg}")
                traceback.print_exc()
            try:
                self.wfile.write(f"data: {json.dumps({'error': error_msg})}\n\n".encode('utf-8'))
                self.wfile.flush()
            except:
                pass  # Client already disconnected


def run_web_ui(model, tokenizer, model_name, system_prompt, args):
    """Run the web UI server."""
    global global_model, global_tokenizer, global_processor, global_args, global_system_prompt, global_model_name
    global conversation_manager
    
    global_model = model
    global_tokenizer = tokenizer
    global_args = args
    global_system_prompt = system_prompt
    global_model_name = model_name
    
    # Sync active_model.local_model with the loaded model
    try:
        from utils.api_chat import get_active_model, save_active_model
        active = get_active_model()
        needs_sync = (
            not active.get('local_model')
            or (active.get('mode') == 'api' and not active.get('api_provider'))
        )
        if needs_sync:
            save_active_model(
                mode='local',
                local_model=model_name,
                api_provider=active.get('api_provider', ''),
                api_model=active.get('api_model', '')
            )
            print(f"[INFO] Reset active_model: mode=local, local_model={model_name}")
    except Exception as e:
        print(f"[Warning] Could not sync active_model config: {e}")
    
    # Load processor for multimodal input
    model_path = find_model_path(args.model)  # Convert model name to actual path
    if model_path:
        try:
            global_processor = AutoProcessor.from_pretrained(model_path)
            print(f"[DEBUG] Loaded processor from {model_path}")
        except Exception as e:
            print(f"[Warning] Could not load processor: {e}")
            global_processor = None
    else:
        print(f"[Warning] Could not find model path for processor")
        global_processor = None
    
    conversation_manager = ConversationManager(LOGS_DIR)
    
    # Warm-up: pre-load conversation list to avoid cold start delay
    _ = conversation_manager.list_conversations()
    
    port = getattr(args, 'port', 8080)
    
    server = ThreadingHTTPServer(('localhost', port), InferenceChatHandler)
    
    url = f"http://localhost:{port}"
    print(f"\n{'='*50}")
    print(f"Inference Chat Web UI")
    print(f"{'='*50}")
    print(f"Model: {model_name}")
    print(f"URL: {url}")
    print(f"Press Ctrl+C to stop")
    print(f"{'='*50}\n")
    
    # Open browser
    webbrowser.open(url)
    
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down server...")
        server.shutdown()


def run_inference(args):
    """
    Load model and start chat.
    """
    if not torch.cuda.is_available():
        print("Error: CUDA is required for inference. Please run on a GPU-enabled machine.")
        return

    # Find model path from name
    model_path = find_model_path(args.model)
    
    if model_path is None:
        print(f"Error: Model '{args.model}' not found.")
        print(f"Searched in: {MODEL_BASE_DIR}/, {MODEL_BASE_DIR}/train/")
        print("\nUse --help_model to see available models.")
        return
    
    model_name = get_model_name(model_path)
    
    # Infer model_type from model name if not explicitly set or using default
    # e.g., "train/ministral_3_3b_instruct-xxx" -> "ministral_3_3b_instruct"
    # e.g., "ministral_3_3b_instruct" -> "ministral_3_3b_instruct"
    model_base_name = os.path.basename(args.model)
    
    # Dynamically get available architectures from architectures/ folder
    arch_dir = os.path.join(os.path.dirname(__file__), "architectures")
    available_archs = []
    if os.path.exists(arch_dir):
        for f in os.listdir(arch_dir):
            if f.endswith(".py") and not f.startswith("_"):
                available_archs.append(f[:-3])  # Remove .py extension
    
    # Sort by length descending to match longer names first (e.g., mHC variant before base)
    available_archs.sort(key=len, reverse=True)
    
    # Try to infer model_type from model name
    inferred_type = None
    for arch_name in available_archs:
        if model_base_name.startswith(arch_name):
            inferred_type = arch_name
            break
    
    if inferred_type:
        args.model_type = inferred_type
    else:
        print(f"Error: Could not infer model_type from '{model_base_name}'")
        print(f"Available architectures: {', '.join(available_archs)}")
        print(f"Use --model_type to specify explicitly.")
        return
    
    if args.debug:
        print(f"[DEBUG] Using model_type: {args.model_type}")
    
    # Load system prompt
    system_prompt = load_system_prompt(model_path, args.model_type)
    print(f"System prompt loaded ({len(system_prompt)} chars)")
    
    print(f"Loading tokenizer from {model_path}...")
    tokenizer = AutoTokenizer.from_pretrained(model_path, trust_remote_code=True)
    set_chat_template_from_file(tokenizer, model_path, args.model_type, debug=args.debug)
    if tokenizer.pad_token is None and tokenizer.eos_token is not None:
        tokenizer.pad_token = tokenizer.eos_token

    # Validate and set KV cache dtype
    kv_cache_dtype = getattr(args, 'kv_cache_dtype', 'bf16')
    if kv_cache_dtype == 'fp8':
        # Check FP8 support (requires CUDA compute capability 8.9+ and PyTorch 2.1+)
        try:
            if not torch.cuda.is_available():
                print(f"[Warning] CUDA not available, falling back to bf16 for KV cache")
                kv_cache_dtype = 'bf16'
            else:
                # Check compute capability
                cc = torch.cuda.get_device_capability()
                if cc[0] < 8 or (cc[0] == 8 and cc[1] < 9):
                    print(f"[Warning] FP8 requires compute capability 8.9+, current: {cc[0]}.{cc[1]}")
                    print(f"[Warning] Falling back to bf16 for KV cache")
                    kv_cache_dtype = 'bf16'
                else:
                    # Check if float8_e4m3fn is available
                    try:
                        _ = torch.float8_e4m3fn
                        print(f"[Info] FP8 KV cache enabled (compute capability {cc[0]}.{cc[1]})")
                    except AttributeError:
                        print(f"[Warning] PyTorch version does not support FP8, falling back to bf16")
                        kv_cache_dtype = 'bf16'
        except Exception as e:
            print(f"[Warning] FP8 check failed: {e}, falling back to bf16")
            kv_cache_dtype = 'bf16'
    
    # Store validated dtype in args
    args.kv_cache_dtype = kv_cache_dtype
    
    # Set global KV cache dtype for architecture modules
    import architectures.ministral_3_3b_instruct as arch_module
    arch_module.KV_CACHE_DTYPE = kv_cache_dtype
    
    # Also set for heteroscedastic variant if available
    try:
        import architectures.ministral_3_3b_instruct_heteroscedastic_uncertainty as hetero_module
        hetero_module.KV_CACHE_DTYPE = kv_cache_dtype
    except ImportError:
        pass
    
    print(f"Loading model from {model_path}...")
    print(f"KV cache dtype: {kv_cache_dtype}")
    # Set model_path for get_model to use
    args.model_path = model_path
    model = model_module.get_model(args)
    print(f"Model loaded on {next(model.parameters()).device}")
    model.eval()
    
    # Wrap with DataParallel if multiple GPUs available
    if torch.cuda.device_count() > 1:
        print(f"[INFO] Using DataParallel with {torch.cuda.device_count()} GPUs")
        model = nn.DataParallel(model)
    
    # Set up tool adapter for the model type
    if TOOLS_AVAILABLE:
        set_adapter(args.model_type)
    
    # Start chat or web UI
    if getattr(args, 'cli', False):
        run_chat(model, tokenizer, model_name, system_prompt, args)
    else:
        run_web_ui(model, tokenizer, model_name, system_prompt, args)


if __name__ == "__main__":
    # Check for detailed help (--arg_name --help)
    from utils.detailed_help import check_detailed_help
    check_detailed_help()
    
    parser = argparse.ArgumentParser(description="Interactive chat with LLM")
    
    # Help for models
    parser.add_argument("--help_model", action="store_true",
                        help="Show available models and exit")
    
    # Model name (not path)
    parser.add_argument("--model", type=str, default=None,
                        help="Model name (e.g., ministral_3_3b_instruct or train/xxx)")
    
    # Generation parameters
    parser.add_argument("--max_length", type=int, default=32768, help="Max new tokens to generate")
    parser.add_argument("--max_context", type=int, default=32768, help="Max context tokens (older tokens truncated). Max 262144 (256k)")
    parser.add_argument("--temperature", type=float, default=0.6, help="Sampling temperature (default: 0.6)")
    parser.add_argument("--top_k", type=int, default=50, help="Top-k sampling")
    parser.add_argument("--kv_cache_dtype", type=str, default="fp8", choices=["bf16", "fp16", "fp8"],
                        help="KV cache dtype for memory optimization (default: fp8, auto-fallback to bf16 if unsupported). FP8 requires CUDA compute capability 8.9+")
    
    # Mode selection
    parser.add_argument("--cli", action="store_true",
                        help="Use terminal CLI mode instead of web UI")
    parser.add_argument("--port", type=int, default=8080,
                        help="Port for web UI server (default: 8080)")

    # Debug options
    parser.add_argument("--debug", action="store_true", help="Enable debug output for detailed logging")
    parser.add_argument("--debug_prompt", action="store_true", help="Print prompt and token ids")
    parser.add_argument("--debug_prompt_tokens", type=int, default=80, help="Prompt token ids to show")
    
    # Refusal Mechanism (enabled by default)
    parser.add_argument("--no_refusal", action="store_true",
                        help="Disable refusal mechanism")
    parser.add_argument("--refusal_threshold", type=float, default=3.0,
                        help="Uncertainty threshold for refusal (std of logits, default: 3.0)")
    parser.add_argument("--refusal_max_retries", type=int, default=3,
                        help="Max retries per token when refused (default: 3)")
    parser.add_argument("--refusal_temp_decay", type=float, default=0.8,
                        help="Temperature multiplier on each retry (default: 0.8)")
    parser.add_argument("--refusal_min_temp", type=float, default=0.1,
                        help="Minimum temperature for refusal mechanism (default: 0.1)")
    parser.add_argument("--refusal_recovery_tokens", type=int, default=3,
                        help="Number of tokens to recover to original temperature after refusal (default: 3)")
    parser.add_argument("--refusal_recovery_method", type=str, default="exponential",
                        choices=["linear", "exponential", "ease_out", "ease_in_out", "step"],
                        help="Temperature recovery curve after refusal (default: exponential)")
    parser.add_argument("--random_seed", type=int, default=-1,
                        help="Random seed (-1 for random, positive for fixed seed)")
    
    # Environment selection
    parser.add_argument("--local", action="store_true",
                        help="Use local paths instead of server paths (default: server). Also sets GPU default to 0.")
    parser.add_argument("--gpu", type=str, default=None,
                        help="GPU IDs to use (e.g., '0' or '6,7'). Default: '6,7' (server) or '0' (local)")
    
    # Add all model arguments from model module
    model_module.add_model_args(parser)

    args = parser.parse_args()
    
    # Set environment based on --local flag
    set_local_mode(args.local)
    MODEL_BASE_DIR = get_model_dir()
    ensure_dirs()
    
    # Handle --help_model
    if args.help_model:
        print_available_models()
        sys.exit(0)
    
    # Require --model if not --help_model
    if args.model is None:
        print("Error: --model is required.")
        print("Use --help_model to see available models.")
        sys.exit(1)
    
    run_inference(args)
